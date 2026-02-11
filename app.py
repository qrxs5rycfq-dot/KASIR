from datetime import datetime, timedelta, timezone


def utc_now():
    """Return current UTC time (timezone-aware)."""
    return datetime.now(timezone.utc)
from functools import wraps
import os
import json
import hmac
import random
import hashlib
import qrcode
from io import BytesIO
import base64

from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect, CSRFError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename

from config import config
from models import db, User, Role, Permission, Category, MenuItem, Table, Order, OrderItem, Payment, Income, Setting, Cart, CartItem, Discount, PendingPrint, Notification, Branch, BranchMenuStock, CashierShift, Expense, ExternalOrder, WebhookLog, City, Brand

# Import USB printer module
try:
    from usb_printer import usb_printer, USBPrinterManager
    USB_PRINTING_AVAILABLE = USBPrinterManager.is_available()
except ImportError:
    usb_printer = None
    USB_PRINTING_AVAILABLE = False

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(config['development'])

# Initialize CSRF Protection
csrf = CSRFProtect(app)

# Initialize Rate Limiter for brute force protection
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Custom Jinja2 filters
def format_number_filter(value):
    """Format angka dengan pemisah ribuan"""
    try:
        return f"{int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return value

def format_currency(value):
    """Format sebagai mata uang Rupiah"""
    try:
        return f"Rp {int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['format_number'] = format_number_filter
app.jinja_env.filters['format_currency'] = format_currency

# Context processor to make config available in all templates
@app.context_processor
def inject_config():
    ctx = {
        'config': {
            'MIDTRANS_CLIENT_KEY': app.config.get('MIDTRANS_CLIENT_KEY', 'SB-Mid-client-XXXXXX'),
            'MIDTRANS_IS_PRODUCTION': app.config.get('MIDTRANS_IS_PRODUCTION', False),
            'APP_NAME': 'Dapoer Teras Obor'
        }
    }
    # Inject current branch info for sidebar
    if current_user.is_authenticated:
        if current_user.branch_id:
            ctx['current_branch'] = current_user.branch
        else:
            ctx['current_branch'] = None  # Admin/owner sees all
        ctx['all_branches'] = Branch.query.filter_by(is_active=True).all()
        ctx['all_cities'] = City.query.filter_by(is_active=True).order_by(City.name).all()
        ctx['all_brands'] = Brand.query.filter_by(is_active=True).order_by(Brand.name).all()
        ctx['is_owner'] = current_user.branch_id is None and current_user.has_role('admin')
    return ctx

def get_user_branch_id():
    """Get the current user's branch_id. Returns None for admin/owner (sees all)."""
    if not current_user.is_authenticated:
        return None
    return current_user.branch_id

def get_default_branch_id():
    """Get branch_id for data creation. Returns user's branch or Pusat branch for admin/owner.
    Ensures records always have a valid branch_id."""
    bid = get_user_branch_id()
    if bid is not None:
        return bid
    # Admin/owner: default to Pusat branch
    pusat = Branch.query.filter_by(code='PUSAT').first()
    return pusat.id if pusat else None

def branch_filter(query, model):
    """Apply branch filter to a query. Admin/owner (branch_id=NULL) sees all data."""
    bid = get_user_branch_id()
    if bid is not None:
        return query.filter(model.branch_id == bid)
    return query

def get_branch_stock(menu_item_id, branch_id):
    """Get BranchMenuStock for a menu item at a specific branch. Creates default if missing."""
    if branch_id is None:
        return None
    bms = BranchMenuStock.query.filter_by(branch_id=branch_id, menu_item_id=menu_item_id).first()
    if not bms:
        bms = BranchMenuStock(branch_id=branch_id, menu_item_id=menu_item_id, stock=100, is_available=True)
        db.session.add(bms)
        db.session.flush()
    return bms

def get_menu_with_branch_stock(menu_items, branch_id):
    """Attach per-branch stock/availability to menu item dicts. Returns list of dicts."""
    if branch_id is None:
        # Owner sees global data â€“ use MenuItem's own stock as fallback
        return [item.to_dict() for item in menu_items]
    
    # Batch-load all branch stock for this branch
    item_ids = [item.id for item in menu_items]
    stocks = {bms.menu_item_id: bms for bms in
              BranchMenuStock.query.filter(
                  BranchMenuStock.branch_id == branch_id,
                  BranchMenuStock.menu_item_id.in_(item_ids)
              ).all()} if item_ids else {}
    
    result = []
    for item in menu_items:
        d = item.to_dict()
        bms = stocks.get(item.id)
        if bms:
            d['stock'] = bms.stock
            d['is_available'] = bms.is_available
        else:
            d['stock'] = 100  # default
            d['is_available'] = True
        result.append(d)
    return result

# Permission decorator
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_permission(permission):
                flash('Anda tidak memiliki akses ke halaman ini.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not any(current_user.has_role(role) for role in roles):
                flash('Anda tidak memiliki akses ke halaman ini.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Prevent caching for dynamic pages and add security headers
@app.after_request
def add_header(response):
    """Add headers to prevent caching for HTML pages and security headers"""
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    
    # Security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    return response

# CSRF Error handler
@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': False, 'error': 'CSRF token missing or invalid'}), 400
    flash('Sesi telah berakhir. Silakan coba lagi.', 'danger')
    return redirect(request.referrer or url_for('dashboard'))

# Rate limit error handler
@app.errorhandler(429)
def ratelimit_handler(e):
    """Custom handler for rate limit exceeded"""
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': False, 
            'error': 'Terlalu banyak permintaan. Silakan tunggu sebentar.'
        }), 429
    
    flash('Terlalu banyak permintaan. Silakan tunggu 1 menit dan coba lagi.', 'warning')
    return redirect(request.referrer or url_for('login'))

# Initialize database and seed data
def run_migrations():
    """Run database migrations for existing databases"""
    from sqlalchemy import inspect, text
    
    inspector = inspect(db.engine)
    
    # Check if payments table exists and add snap_token column if missing
    if 'payments' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('payments')]
        if 'snap_token' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE payments ADD COLUMN snap_token VARCHAR(255)'))
                conn.commit()
            print("Added snap_token column to payments table")
    
    # Check if order_items table exists and add item_status column if missing
    if 'order_items' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('order_items')]
        if 'item_status' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE order_items ADD COLUMN item_status VARCHAR(20) DEFAULT 'pending'"))
                conn.commit()
            print("Added item_status column to order_items table")
    
    # Check if users table exists and add printer columns if missing
    if 'users' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('users')]
        if 'printer_name' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN printer_name VARCHAR(100)"))
                conn.commit()
            print("Added printer_name column to users table")
        if 'printer_id' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN printer_id VARCHAR(100)"))
                conn.commit()
            print("Added printer_id column to users table")
        if 'force_password_change' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN force_password_change BOOLEAN DEFAULT 0"))
                conn.commit()
            print("Added force_password_change column to users table")
    
    # Check if cart table exists (new feature)
    if 'cart' not in inspector.get_table_names():
        # db.create_all will handle this
        pass
    
    if 'cart_item' not in inspector.get_table_names():
        # db.create_all will handle this
        pass
    
    # Add branch_id columns to existing tables for multi-branch support
    branch_tables = {
        'users': 'branch_id',
        'categories': 'branch_id',
        'menu_items': 'branch_id',
        'tables': 'branch_id',
        'orders': 'branch_id',
        'carts': 'branch_id',
        'discounts': 'branch_id',
        'incomes': 'branch_id',
        'expenses': 'branch_id',
        'cashier_shifts': 'branch_id',
        'notifications': 'branch_id',
        'pending_prints': 'branch_id',
    }
    for tbl_name, col_name in branch_tables.items():
        if tbl_name in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns(tbl_name)]
            if col_name not in columns:
                with db.engine.connect() as conn:
                    conn.execute(text(f"ALTER TABLE {tbl_name} ADD COLUMN {col_name} INTEGER REFERENCES branches(id)"))
                    conn.commit()
                print(f"Added {col_name} column to {tbl_name} table")
    
    # Add source column to orders table for tracking order origin
    if 'orders' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('orders')]
        if 'source' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE orders ADD COLUMN source VARCHAR(30) DEFAULT 'pos'"))
                conn.commit()
            print("Added source column to orders table")
    
    # Remove unique constraint on tables.number if it exists (now scoped per branch)
    if 'tables' in inspector.get_table_names():
        try:
            with db.engine.connect() as conn:
                # SQLite doesn't support DROP CONSTRAINT directly, but create_all handles new schema
                pass
        except Exception:
            pass
    
    # Add city_id and brand_id columns to branches table for multi-outlet hierarchy
    table_names = inspector.get_table_names()
    if 'branches' in table_names and 'cities' in table_names and 'brands' in table_names:
        columns = [col['name'] for col in inspector.get_columns('branches')]
        if 'city_id' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE branches ADD COLUMN city_id INTEGER REFERENCES cities(id)"))
                conn.commit()
            print("Added city_id column to branches table")
        if 'brand_id' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE branches ADD COLUMN brand_id INTEGER REFERENCES brands(id)"))
                conn.commit()
            print("Added brand_id column to branches table")
    
    # Fix legacy data: assign NULL branch_id records to Pusat branch
    if 'branches' in table_names:
        # Table names are from a hardcoded allowlist (not user input), safe for f-string
        allowed_tables = {'orders', 'carts', 'expenses', 'cashier_shifts', 'pending_prints', 'notifications', 'discounts', 'incomes'}
        with db.engine.connect() as conn:
            result = conn.execute(text("SELECT id FROM branches WHERE code = 'PUSAT' LIMIT 1"))
            row = result.fetchone()
            if row:
                pusat_id = row[0]
                for tbl in allowed_tables:
                    if tbl in table_names:
                        cols = [c['name'] for c in inspector.get_columns(tbl)]
                        if 'branch_id' in cols:
                            result = conn.execute(text(f"UPDATE {tbl} SET branch_id = :bid WHERE branch_id IS NULL"), {'bid': pusat_id})
                            if result.rowcount > 0:
                                print(f"Assigned {result.rowcount} {tbl} records with NULL branch_id to Pusat branch")
                conn.commit()

def init_db():
    with app.app_context():
        db.create_all()
        
        # Run migrations for existing databases
        run_migrations()
        
        # Create default permissions
        permissions_data = [
            ('view_dashboard', 'Dapat melihat dashboard'),
            ('manage_orders', 'Dapat mengelola pesanan'),
            ('manage_menu', 'Dapat mengelola menu'),
            ('manage_users', 'Dapat mengelola pengguna'),
            ('manage_tables', 'Dapat mengelola meja'),
            ('view_reports', 'Dapat melihat laporan'),
            ('manage_settings', 'Dapat mengelola pengaturan'),
            ('process_payment', 'Dapat memproses pembayaran'),
            ('view_income', 'Dapat melihat penghasilan'),
            ('manage_income', 'Dapat mengelola penghasilan'),
        ]
        
        for perm_name, perm_desc in permissions_data:
            if not Permission.query.filter_by(name=perm_name).first():
                perm = Permission(name=perm_name, description=perm_desc)
                db.session.add(perm)
        
        db.session.commit()
        
        # Create default roles
        roles_data = {
            'admin': {
                'description': 'Administrator dengan akses penuh',
                'permissions': [p[0] for p in permissions_data]
            },
            'manager': {
                'description': 'Manager dengan akses laporan dan manajemen',
                'permissions': ['view_dashboard', 'manage_orders', 'manage_menu', 'view_reports', 'manage_tables', 'process_payment', 'view_income']
            },
            'kasir': {
                'description': 'Kasir untuk proses pembayaran',
                'permissions': ['view_dashboard', 'manage_orders', 'process_payment']
            },
            'koki': {
                'description': 'Koki untuk mengelola pesanan di dapur',
                'permissions': ['view_dashboard', 'manage_orders']
            },
            'customer': {
                'description': 'Pelanggan untuk memesan online',
                'permissions': ['view_dashboard']
            }
        }
        
        for role_name, role_data in roles_data.items():
            role = Role.query.filter_by(name=role_name).first()
            if not role:
                role = Role(name=role_name, description=role_data['description'])
                db.session.add(role)
                db.session.commit()
            
            # Add permissions to role
            for perm_name in role_data['permissions']:
                perm = Permission.query.filter_by(name=perm_name).first()
                if perm and perm not in role.permissions:
                    role.permissions.append(perm)
        
        db.session.commit()
        
        # Create default city
        default_city = City.query.filter_by(code='PUSAT').first()
        if not default_city:
            default_city = City(
                name='Kota Pusat',
                code='PUSAT',
                is_active=True
            )
            db.session.add(default_city)
            db.session.commit()
        
        # Create default brand
        default_brand = Brand.query.filter_by(code='DTO').first()
        if not default_brand:
            default_brand = Brand(
                name='Dapoer Teras Obor',
                code='DTO',
                description='Brand utama restoran',
                is_active=True
            )
            db.session.add(default_brand)
            db.session.commit()
        
        # Create default branch (Pusat / HQ)
        default_branch = Branch.query.filter_by(code='PUSAT').first()
        if not default_branch:
            default_branch = Branch(
                name='Cabang Pusat',
                code='PUSAT',
                address='Alamat cabang pusat',
                is_active=True,
                city_id=default_city.id,
                brand_id=default_brand.id
            )
            db.session.add(default_branch)
            db.session.commit()
        else:
            # Assign city and brand to existing Pusat branch if missing
            if default_branch.city_id is None:
                default_branch.city_id = default_city.id
            if default_branch.brand_id is None:
                default_branch.brand_id = default_brand.id
            db.session.commit()
        
        # Create default admin user (owner - no branch = sees all)
        if not User.query.filter_by(username='admin').first():
            admin_role = Role.query.filter_by(name='admin').first()
            admin = User(
                username='admin',
                email='admin@kasir.com',
                full_name='Owner / Administrator',
                force_password_change=True,
                branch_id=None  # Owner sees all branches
            )
            admin.set_password('admin123')
            admin.roles.append(admin_role)
            db.session.add(admin)
        
        # Create default kasir user (assigned to Pusat branch)
        if not User.query.filter_by(username='kasir').first():
            kasir_role = Role.query.filter_by(name='kasir').first()
            kasir = User(
                username='kasir',
                email='kasir@kasir.com',
                full_name='Kasir Utama',
                force_password_change=True,
                branch_id=default_branch.id
            )
            kasir.set_password('kasir123')
            kasir.roles.append(kasir_role)
            db.session.add(kasir)
        
        # Create default koki user (assigned to Pusat branch)
        if not User.query.filter_by(username='koki').first():
            koki_role = Role.query.filter_by(name='koki').first()
            koki = User(
                username='koki',
                email='koki@kasir.com',
                full_name='Koki Dapur',
                force_password_change=True,
                branch_id=default_branch.id
            )
            koki.set_password('koki123')
            koki.roles.append(koki_role)
            db.session.add(koki)
        
        db.session.commit()
        
        # Create default categories
        categories_data = [
            ('Nasi Goreng', 'Menu nasi goreng berbagai varian', 'fa-bowl-rice', 1),
            ('Mie', 'Menu mie berbagai varian', 'fa-utensils', 2),
            ('Kwetiau', 'Menu kwetiau berbagai varian', 'fa-plate-wheat', 3),
            ('Menu Lain', 'Menu lainnya', 'fa-drumstick-bite', 4),
            ('Paket', 'Menu paket hemat', 'fa-box', 5),
            ('Snack', 'Makanan ringan', 'fa-cookie', 6),
            ('Minuman', 'Berbagai minuman segar', 'fa-mug-hot', 7),
        ]
        
        for cat_name, cat_desc, cat_icon, cat_order in categories_data:
            if not Category.query.filter_by(name=cat_name).first():
                cat = Category(name=cat_name, description=cat_desc, icon=cat_icon, order=cat_order, branch_id=default_branch.id)
                db.session.add(cat)
        
        db.session.commit()
        
        # Create menu items from PDF menu (Solaria style)
        seed_menu_items(default_branch.id)
        
        # Create default tables
        for i in range(1, 21):
            table_num = f"{i:02d}"
            if not Table.query.filter_by(number=table_num, branch_id=default_branch.id).first():
                table = Table(
                    number=table_num,
                    name=f"Meja {i}",
                    capacity=4 if i <= 15 else 6,
                    branch_id=default_branch.id
                )
                db.session.add(table)
        
        db.session.commit()
        
        # Create BranchMenuStock entries for default branch
        all_menu_items = MenuItem.query.all()
        for mi in all_menu_items:
            if not BranchMenuStock.query.filter_by(branch_id=default_branch.id, menu_item_id=mi.id).first():
                bms = BranchMenuStock(branch_id=default_branch.id, menu_item_id=mi.id, stock=100, is_available=True)
                db.session.add(bms)
        
        db.session.commit()
        print("Database initialized successfully!")

def seed_menu_items(branch_id=None):
    """Seed menu items from Solaria menu PDF"""
    
    # Get categories
    nasi_goreng = Category.query.filter_by(name='Nasi Goreng').first()
    mie = Category.query.filter_by(name='Mie').first()
    kwetiau = Category.query.filter_by(name='Kwetiau').first()
    menu_lain = Category.query.filter_by(name='Menu Lain').first()
    paket = Category.query.filter_by(name='Paket').first()
    snack = Category.query.filter_by(name='Snack').first()
    minuman = Category.query.filter_by(name='Minuman').first()
    
    menu_items_data = [
        # Nasi Goreng
        {'code': '111', 'name': 'Nasi Goreng Mlarat', 'price': 20000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '121', 'name': 'Nasi Goreng Spesial', 'price': 22000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '131', 'name': 'Nasi Goreng Cabe Ijo', 'price': 22000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=400&h=300&fit=crop'},
        {'code': '141', 'name': 'Nasi Goreng Sosis', 'price': 23000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1596560548464-f010549b84d7?w=400&h=300&fit=crop'},
        {'code': '151', 'name': 'Nasi Goreng Modern Warno', 'price': 24000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '161', 'name': 'Nasi Goreng Terimaskenthir', 'price': 25000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '171', 'name': 'Nasi Goreng Pete', 'price': 25000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1569058242253-92a9c755a0ec?w=400&h=300&fit=crop'},
        {'code': '181', 'name': 'Nasi Goreng Seafood', 'price': 28000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        
        # Mie
        {'code': '212', 'name': 'Mie Goreng Ayam', 'price': 22000, 'category': mie, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '222', 'name': 'Mie Siram Ayam', 'price': 22000, 'category': mie, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1555126634-323283e090fa?w=400&h=300&fit=crop'},
        {'code': '232', 'name': 'Mie Goreng Seafood', 'price': 28000, 'category': mie, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '242', 'name': 'Mie Siram Seafood', 'price': 28000, 'category': mie, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=400&h=300&fit=crop'},
        {'code': '252', 'name': 'Mie Goreng Sapi', 'price': 30000, 'category': mie, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '262', 'name': 'Mie Siram Sapi', 'price': 30000, 'category': mie, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=400&h=300&fit=crop'},
        
        # Kwetiau
        {'code': '414', 'name': 'Kwetiau Ayam Goreng', 'price': 25000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '424', 'name': 'Kwetiau Ayam Siram', 'price': 25000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '434', 'name': 'Kwetiau Seafood Goreng', 'price': 28000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '444', 'name': 'Kwetiau Seafood Siram', 'price': 28000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=400&h=300&fit=crop'},
        {'code': '454', 'name': 'Kwetiau Sapi Goreng', 'price': 30000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '464', 'name': 'Kwetiau Sapi Siram', 'price': 30000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        
        # Menu Lain
        {'code': '515', 'name': 'Cap Cay Goreng Ayam', 'price': 23000, 'category': menu_lain, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '525', 'name': 'Cap Cay Goreng Seafood', 'price': 28000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=400&h=300&fit=crop'},
        {'code': '535', 'name': 'Sapo Tahu Ayam', 'price': 27000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=400&h=300&fit=crop'},
        {'code': '545', 'name': 'Sapo Tahu Seafood', 'price': 30000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '555', 'name': 'Nasi Putih', 'price': 5000, 'category': menu_lain, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1516684732162-798a0062be99?w=400&h=300&fit=crop'},
        {'code': '565', 'name': 'Telur Mata Sapi / Dadar', 'price': 5000, 'category': menu_lain, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1510693206972-df098062cb71?w=400&h=300&fit=crop'},
        
        # Snack
        {'code': '313', 'name': 'Fish Cake', 'price': 12000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '323', 'name': 'Kentang Goreng', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        {'code': '333', 'name': 'Otak Otak', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '343', 'name': 'Sosis Goreng', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '353', 'name': 'Sosis Bakar', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '363', 'name': 'Mix OTP', 'price': 20000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        
        # Paket
        {'code': '616', 'name': 'Nasi Goreng Cabe Ijo + Teh', 'price': 25000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '626', 'name': 'Kwetiau Ayam Goreng + Teh', 'price': 28000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '636', 'name': 'Nasi Goreng Spesial + Lemon Tea', 'price': 33000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '646', 'name': 'Kwetiau Ayam Goreng + Thai Tea', 'price': 35000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '656', 'name': '2 Thai Tea + Kentang Goreng', 'price': 38000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        {'code': '666', 'name': '2 Cappucino + Mix OTP', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=400&h=300&fit=crop'},
        {'code': '676', 'name': 'Nasi Goreng + Kwetiau Seafood + Blackcurant', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '686', 'name': 'Nasi + Sapo Tahu Seafood + Lemonade', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=400&h=300&fit=crop'},
        
        # Minuman
        {'code': '717', 'name': 'Teh Mlarat', 'price': 3000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1576092768241-dec231879fc3?w=400&h=300&fit=crop'},
        {'code': '727', 'name': 'Teh Manis', 'price': 5000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1576092768241-dec231879fc3?w=400&h=300&fit=crop'},
        {'code': '737', 'name': 'Air Mineral', 'price': 5000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': False, 'image': 'https://images.unsplash.com/photo-1523362628745-0c100150b504?w=400&h=300&fit=crop'},
        {'code': '747', 'name': 'Kopi Hitam', 'price': 6000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1514432324607-a09d9b4aefdd?w=400&h=300&fit=crop'},
        {'code': '757', 'name': 'Green Tea', 'price': 13000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1556679343-c7306c1976bc?w=400&h=300&fit=crop'},
        {'code': '767', 'name': 'Thai Tea', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1558857563-b371033873b8?w=400&h=300&fit=crop'},
        {'code': '777', 'name': 'Green Tea Milk', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1515823064-d6e0c04616a7?w=400&h=300&fit=crop'},
        {'code': '787', 'name': 'Milo', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1517578239113-b03992dcdd25?w=400&h=300&fit=crop'},
        {'code': '797', 'name': 'Thai Tea Milo', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1558857563-b371033873b8?w=400&h=300&fit=crop'},
        {'code': '708', 'name': 'Cappucino', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=400&h=300&fit=crop'},
        {'code': '718', 'name': 'Teh Tarik', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1571934811356-5cc061b6821f?w=400&h=300&fit=crop'},
        {'code': '728', 'name': 'Lemon Tea', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1556679343-c7306c1976bc?w=400&h=300&fit=crop'},
        {'code': '738', 'name': 'Lemonade', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1621263764928-df1444c5e859?w=400&h=300&fit=crop'},
        {'code': '748', 'name': 'Blackcurrant', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1544145945-f90425340c7e?w=400&h=300&fit=crop'},
    ]
    
    for item_data in menu_items_data:
        existing_item = MenuItem.query.filter_by(code=item_data['code']).first()
        if not existing_item:
            menu_item = MenuItem(
                code=item_data['code'],
                name=item_data['name'],
                price=item_data['price'],
                category_id=item_data['category'].id if item_data['category'] else None,
                has_spicy_option=item_data.get('has_spicy', False),
                has_temperature_option=item_data.get('has_temp', False),
                is_popular=item_data.get('popular', False),
                image=item_data.get('image', ''),
                description=f"Menu {item_data['name']} yang lezat",
                branch_id=branch_id
            )
            db.session.add(menu_item)
        else:
            # Update image URL if it changed
            existing_item.image = item_data.get('image', existing_item.image)
    
    db.session.commit()

# Generate QR Code for table
def generate_table_qr(table_number):
    app_url = app.config.get('APP_URL', 'http://localhost:8000')
    order_url = f"{app_url}/order/online/{table_number}"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(order_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to file
    qr_folder = app.config.get('QR_CODE_FOLDER', 'static/qrcodes')
    os.makedirs(qr_folder, exist_ok=True)
    qr_path = os.path.join(qr_folder, f"table_{table_number}.png")
    img.save(qr_path)
    
    # Also return base64 for display
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    
    return qr_path, img_str

# ==================== ROUTES ====================

# Public routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute", methods=["POST"])  # Only limit POST (actual login attempts)
def login():
    if current_user.is_authenticated:
        # Check if force password change is required
        if hasattr(current_user, 'force_password_change') and current_user.force_password_change:
            return redirect(url_for('change_password'))
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Akun Anda telah dinonaktifkan. Hubungi administrator.', 'danger')
                return render_template('auth/login.html')
            
            # Check if user's branch is active (branch users only)
            if user.branch_id:
                user_branch = db.session.get(Branch, user.branch_id)
                if not user_branch:
                    flash('Cabang yang ditugaskan tidak ditemukan. Hubungi owner/admin pusat.', 'danger')
                    return render_template('auth/login.html')
                if not user_branch.is_active:
                    flash(f'Cabang "{user_branch.name}" sedang nonaktif. Hubungi owner/admin pusat.', 'danger')
                    return render_template('auth/login.html')
            
            login_user(user, remember=remember)
            user.last_login = utc_now()
            db.session.commit()
            
            # Check if force password change is required
            if hasattr(user, 'force_password_change') and user.force_password_change:
                flash('Silakan ganti password default Anda untuk keamanan.', 'warning')
                return redirect(url_for('change_password'))
            
            next_page = request.args.get('next')
            flash(f'Selamat datang, {user.full_name or user.username}!', 'success')
            
            # Redirect koki to kitchen display
            if user.has_role('koki') and not user.has_role('admin'):
                return redirect(next_page or url_for('kitchen'))
            
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Username atau password salah!', 'danger')
    
    return render_template('auth/login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        
        # Validation
        if User.query.filter_by(username=username).first():
            flash('Username sudah digunakan!', 'danger')
            return render_template('auth/register.html')
        
        if User.query.filter_by(email=email).first():
            flash('Email sudah digunakan!', 'danger')
            return render_template('auth/register.html')
        
        if password != confirm_password:
            flash('Password tidak cocok!', 'danger')
            return render_template('auth/register.html')
        
        if len(password) < 6:
            flash('Password minimal 6 karakter!', 'danger')
            return render_template('auth/register.html')
        
        # Create user
        customer_role = Role.query.filter_by(name='customer').first()
        user = User(
            username=username,
            email=email,
            full_name=full_name
        )
        user.set_password(password)
        if customer_role:
            user.roles.append(customer_role)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registrasi berhasil! Silakan login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('auth/register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Force password change page for first-time login or security requirements"""
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Validate current password
        if not current_user.check_password(current_password):
            flash('Password saat ini salah!', 'danger')
            return render_template('auth/change_password.html')
        
        # Validate new password
        if len(new_password) < 8:
            flash('Password baru minimal 8 karakter!', 'danger')
            return render_template('auth/change_password.html')
        
        if new_password != confirm_password:
            flash('Password baru tidak cocok!', 'danger')
            return render_template('auth/change_password.html')
        
        # Check password strength (at least 1 uppercase, 1 lowercase, 1 number)
        import re
        if not re.search(r'[A-Z]', new_password):
            flash('Password harus mengandung minimal 1 huruf besar!', 'danger')
            return render_template('auth/change_password.html')
        if not re.search(r'[a-z]', new_password):
            flash('Password harus mengandung minimal 1 huruf kecil!', 'danger')
            return render_template('auth/change_password.html')
        if not re.search(r'[0-9]', new_password):
            flash('Password harus mengandung minimal 1 angka!', 'danger')
            return render_template('auth/change_password.html')
        
        # Update password
        current_user.set_password(new_password)
        current_user.force_password_change = False
        db.session.commit()
        
        flash('Password berhasil diubah!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('auth/change_password.html')


# Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    # Get only completed/paid orders today
    today_orders = branch_filter(Order.query, Order).filter(
        db.func.date(Order.created_at) == today
    ).all()
    
    # Calculate income from paid orders only
    paid_orders_today = [o for o in today_orders if o.payment and o.payment.status == 'paid']
    total_income_today = sum(o.total for o in paid_orders_today)
    total_orders_today = len(paid_orders_today)
    
    # Yesterday's data for real growth percentages
    yesterday_orders = branch_filter(Order.query, Order).filter(
        db.func.date(Order.created_at) == yesterday
    ).all()
    paid_orders_yesterday = [o for o in yesterday_orders if o.payment and o.payment.status == 'paid']
    total_income_yesterday = sum(o.total for o in paid_orders_yesterday)
    total_orders_yesterday = len(paid_orders_yesterday)
    
    # Calculate real growth percentages
    def calc_growth(current, previous):
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round(((current - previous) / previous) * 100, 1)
    
    income_growth = calc_growth(total_income_today, total_income_yesterday)
    orders_growth = calc_growth(total_orders_today, total_orders_yesterday)
    
    today_avg = total_income_today // total_orders_today if total_orders_today > 0 else 0
    yesterday_avg = total_income_yesterday // total_orders_yesterday if total_orders_yesterday > 0 else 0
    avg_growth = calc_growth(today_avg, yesterday_avg)
    
    # Get popular items from actual order statistics (last 30 days)
    # Query to find most ordered items
    popular_query = db.session.query(
        OrderItem.menu_item_id,
        db.func.sum(OrderItem.quantity).label('total_ordered')
    ).join(Order).filter(
        Order.created_at >= datetime.now() - timedelta(days=30),
        Order.status.in_(['completed', 'processing'])
    ).group_by(OrderItem.menu_item_id).order_by(
        db.func.sum(OrderItem.quantity).desc()
    ).limit(6).all()
    
    # Get menu items for popular items
    popular_item_ids = [item[0] for item in popular_query if item[0]]
    if popular_item_ids:
        popular_items = MenuItem.query.filter(MenuItem.id.in_(popular_item_ids)).all()
        # Sort by order count
        item_order = {item_id: idx for idx, (item_id, _) in enumerate(popular_query)}
        popular_items.sort(key=lambda x: item_order.get(x.id, 999))
    else:
        # Fallback to marked popular items if no orders yet
        popular_items = MenuItem.query.filter_by(is_popular=True).limit(6).all()
    
    # Get recent orders
    recent_orders = branch_filter(Order.query, Order).order_by(Order.created_at.desc()).limit(10).all()
    
    # Get tables status - based on active orders
    tables = Table.query.filter_by(is_active=True).all()
    
    # Calculate occupied tables from current active orders
    active_orders = branch_filter(Order.query, Order).filter(
        Order.status.in_(['pending', 'processing']),
        Order.table_id.isnot(None)
    ).all()
    occupied_table_ids = set(o.table_id for o in active_orders)
    
    # Update tables status dynamically
    for table in tables:
        if table.id in occupied_table_ids:
            table.status = 'occupied'
        else:
            table.status = 'available'
    
    # Low stock items (stock <= 10) - per-branch
    bid = get_user_branch_id()
    if bid:
        low_stock_items_raw = db.session.query(MenuItem, BranchMenuStock).join(
            BranchMenuStock, BranchMenuStock.menu_item_id == MenuItem.id
        ).filter(
            BranchMenuStock.branch_id == bid,
            BranchMenuStock.is_available == True,
            BranchMenuStock.stock <= 10
        ).order_by(BranchMenuStock.stock.asc()).all()
        # Attach stock to menu item objects for template compatibility
        low_stock_items = []
        for mi, bms in low_stock_items_raw:
            mi._branch_stock = bms.stock
            low_stock_items.append(mi)
    else:
        low_stock_items = MenuItem.query.filter(
            MenuItem.is_available == True,
            MenuItem.stock <= 10
        ).order_by(MenuItem.stock.asc()).all()
        for mi in low_stock_items:
            mi._branch_stock = mi.stock
    
    return render_template('dashboard.html',
                         total_income_today=total_income_today,
                         total_orders_today=total_orders_today,
                         income_growth=income_growth,
                         orders_growth=orders_growth,
                         avg_growth=avg_growth,
                         popular_items=popular_items,
                         recent_orders=recent_orders,
                         tables=tables,
                         low_stock_items=low_stock_items,
                         now=datetime.now())

# POS (Kasir)
@app.route('/pos')
@login_required
@role_required('admin', 'manager', 'kasir')
def pos():
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    menu_items = MenuItem.query.filter_by(is_available=True).all()
    tables = Table.query.filter_by(is_active=True).all()
    
    return render_template('pos.html',
                         categories=categories,
                         menu_items=menu_items,
                         tables=tables,
                         now=datetime.now())

# Online Order (via QR code)
@app.route('/order/online/<table_number>')
def online_order(table_number):
    table = Table.query.filter_by(number=table_number).first()
    if not table:
        flash('Meja tidak ditemukan!', 'danger')
        return redirect(url_for('login'))
    
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    menu_items = MenuItem.query.filter_by(is_available=True).all()
    
    return render_template('online_order.html',
                         table=table,
                         categories=categories,
                         menu_items=menu_items,
                         now=datetime.now())

# API Routes
@app.route('/api/menu')
def api_get_menu():
    bid = get_user_branch_id()
    menu_items = MenuItem.query.filter_by(is_available=True).all() if bid is None else MenuItem.query.all()
    items = get_menu_with_branch_stock(menu_items, bid)
    # Filter to available only (per-branch availability)
    return jsonify([i for i in items if i['is_available']])

@app.route('/api/menu/category/<int:category_id>')
def api_get_menu_by_category(category_id):
    bid = get_user_branch_id()
    base = MenuItem.query.filter_by(category_id=category_id)
    menu_items = base.filter_by(is_available=True).all() if bid is None else base.all()
    items = get_menu_with_branch_stock(menu_items, bid)
    return jsonify([i for i in items if i['is_available']])

# ============================================
# CART API - Database-backed shopping cart
# ============================================

def get_or_create_cart():
    """Get current user's cart or create new one"""
    if current_user.is_authenticated:
        cart = Cart.query.filter_by(user_id=current_user.id).first()
        if not cart:
            cart = Cart(user_id=current_user.id, branch_id=get_default_branch_id())
            db.session.add(cart)
            db.session.commit()
    else:
        # For guest users, use session
        session_id = session.get('cart_session_id')
        if not session_id:
            import uuid
            session_id = str(uuid.uuid4())
            session['cart_session_id'] = session_id
        
        cart = Cart.query.filter_by(session_id=session_id).first()
        if not cart:
            cart = Cart(session_id=session_id)
            db.session.add(cart)
            db.session.commit()
    
    return cart

@app.route('/api/cart')
def api_get_cart():
    """Get current cart items"""
    cart = get_or_create_cart()
    return jsonify({'success': True, 'cart': cart.to_dict()})

@app.route('/api/cart/add', methods=['POST'])
def api_add_to_cart():
    """Add item to cart"""
    try:
        data = request.json
        menu_item_id = data.get('menu_item_id')
        quantity = data.get('quantity', 1)
        spice_level = data.get('spice_level')
        temperature = data.get('temperature')
        notes = data.get('notes', '')
        
        menu_item = db.session.get(MenuItem, menu_item_id)
        if not menu_item:
            return jsonify({'success': False, 'error': 'Menu item not found'}), 404
        
        cart = get_or_create_cart()
        
        # Check if same item with same options exists
        existing_item = CartItem.query.filter_by(
            cart_id=cart.id,
            menu_item_id=menu_item_id,
            spice_level=spice_level,
            temperature=temperature,
            notes=notes
        ).first()
        
        if existing_item:
            existing_item.quantity += quantity
            existing_item.update_subtotal()
        else:
            cart_item = CartItem(
                cart_id=cart.id,
                menu_item_id=menu_item_id,
                name=menu_item.name,
                price=menu_item.price,
                quantity=quantity,
                subtotal=menu_item.price * quantity,
                spice_level=spice_level,
                temperature=temperature,
                notes=notes
            )
            db.session.add(cart_item)
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/update/<int:item_id>', methods=['PUT'])
def api_update_cart_item(item_id):
    """Update cart item quantity"""
    try:
        data = request.json
        quantity = data.get('quantity', 1)
        
        cart = get_or_create_cart()
        cart_item = CartItem.query.filter_by(id=item_id, cart_id=cart.id).first()
        
        if not cart_item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        
        if quantity <= 0:
            db.session.delete(cart_item)
        else:
            cart_item.quantity = quantity
            cart_item.update_subtotal()
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/remove/<int:item_id>', methods=['DELETE'])
def api_remove_cart_item(item_id):
    """Remove item from cart"""
    try:
        cart = get_or_create_cart()
        cart_item = CartItem.query.filter_by(id=item_id, cart_id=cart.id).first()
        
        if not cart_item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        
        db.session.delete(cart_item)
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/clear', methods=['DELETE'])
def api_clear_cart():
    """Clear all items from cart"""
    try:
        cart = get_or_create_cart()
        CartItem.query.filter_by(cart_id=cart.id).delete()
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/settings', methods=['PUT'])
def api_update_cart_settings():
    """Update cart settings (table, order type, customer name)"""
    try:
        data = request.json
        cart = get_or_create_cart()
        
        if 'table_id' in data:
            cart.table_id = data['table_id'] if data['table_id'] else None
        if 'order_type' in data:
            cart.order_type = data['order_type']
        if 'customer_name' in data:
            cart.customer_name = data['customer_name']
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================

@app.route('/api/order', methods=['POST'])
@limiter.limit("60 per minute")  # Protect order creation from abuse
def api_create_order():
    try:
        data = request.json
        items = data.get('items', [])
        table_id = data.get('table_id')
        order_type = data.get('order_type', 'dine_in')
        customer_name = data.get('customer_name', '')
        notes = data.get('notes', '')
        payment_method = data.get('payment_method', 'cash')
        paid_amount = data.get('paid_amount', 0)
        
        if not items:
            return jsonify({'error': 'Keranjang kosong'}), 400
        
        # Generate order number
        order_number = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Create order
        order = Order(
            order_number=order_number,
            user_id=current_user.id if current_user.is_authenticated else None,
            table_id=table_id,
            customer_name=customer_name,
            order_type=order_type,
            notes=notes,
            branch_id=get_default_branch_id()
        )
        db.session.add(order)
        db.session.flush()
        
        # Add order items
        order_branch_id = get_default_branch_id()
        for item in items:
            menu_item = db.session.get(MenuItem, item.get('menu_item_id') or item.get('id'))
            if menu_item:
                qty = item.get('quantity', 1)
                
                # Check stock availability (per-branch)
                if order_branch_id:
                    bms = get_branch_stock(menu_item.id, order_branch_id)
                    if bms.stock < qty:
                        db.session.rollback()
                        return jsonify({'error': f'Stok "{menu_item.name}" tidak cukup (sisa {bms.stock})'}), 400
                else:
                    # Owner/admin fallback: use global stock
                    if menu_item.stock < qty:
                        db.session.rollback()
                        return jsonify({'error': f'Stok "{menu_item.name}" tidak cukup (sisa {menu_item.stock})'}), 400
                
                order_item = OrderItem(
                    order_id=order.id,
                    menu_item_id=menu_item.id,
                    name=menu_item.name,
                    price=menu_item.price,
                    quantity=qty,
                    subtotal=menu_item.price * qty,
                    spice_level=item.get('spice_level'),
                    temperature=item.get('temperature'),
                    notes=item.get('notes')
                )
                db.session.add(order_item)
                
                # Decrement stock (per-branch)
                if order_branch_id:
                    bms = get_branch_stock(menu_item.id, order_branch_id)
                    bms.stock -= qty
                else:
                    menu_item.stock -= qty
        
        # Calculate totals
        order.calculate_totals()
        
        # Apply discount if provided
        discount_amount = data.get('discount', 0)
        discount_code = data.get('discount_code')
        
        if discount_amount > 0 and discount_code:
            # Find and update discount usage
            discount = Discount.query.filter_by(code=discount_code).first()
            if discount:
                discount.usage_count += 1
            
            # Apply discount to order
            order.discount = discount_amount
            order.total = max(0, order.subtotal - discount_amount)
        
        # Create payment record
        final_total = order.total
        is_cash_paid = payment_method == 'cash' and paid_amount >= final_total
        payment = Payment(
            order_id=order.id,
            payment_method=payment_method,
            amount=final_total,
            paid_amount=paid_amount if payment_method == 'cash' else 0,
            change_amount=max(0, paid_amount - final_total) if payment_method == 'cash' else 0,
            status='paid' if is_cash_paid else 'pending'
        )
        
        if is_cash_paid:
            payment.paid_at = utc_now()
            order.status = 'processing'
        
        db.session.add(payment)
        db.session.commit()
        
        # Create notification for new order
        total_formatted = f"{order.total:,}".replace(',', '.')
        create_notification(
            type='order_new',
            title='Pesanan Baru!',
            message=f'Order #{order.order_number} - {customer_name or "Guest"} - Rp {total_formatted}',
            data={'order_id': order.id, 'order_number': order.order_number}
        )
        
        # Update table status
        if table_id:
            table = db.session.get(Table, table_id)
            if table:
                table.status = 'occupied'
                db.session.commit()
        
        # For online payment (Midtrans), generate Snap token
        if payment_method == 'online':
            midtrans_order_id = f"DTO-{order.order_number}"
            payment.midtrans_order_id = midtrans_order_id
            payment.payment_method = 'midtrans'
            
            # Generate Midtrans Snap token using Snap API
            snap_token = generate_midtrans_snap_token(order, midtrans_order_id)
            
            if snap_token:
                payment.snap_token = snap_token
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'order': order.to_dict(),
                    'payment_method': 'online',
                    'snap_token': snap_token,
                    'midtrans_client_key': app.config.get('MIDTRANS_CLIENT_KEY'),
                    'message': 'Pesanan dibuat, silakan lakukan pembayaran.'
                })
            else:
                # Fallback: still create order but mark as pending
                db.session.commit()
                return jsonify({
                    'success': True,
                    'order': order.to_dict(),
                    'payment_method': 'online',
                    'snap_token': None,
                    'error_payment': 'Gagal menghubungi payment gateway, silakan coba lagi.',
                    'message': 'Pesanan dibuat dengan status pending.'
                })
        
        # Clear cart after successful order
        if current_user.is_authenticated:
            cart = Cart.query.filter_by(user_id=current_user.id).first()
        else:
            session_id = session.get('cart_session_id')
            cart = Cart.query.filter_by(session_id=session_id).first() if session_id else None
        
        if cart:
            CartItem.query.filter_by(cart_id=cart.id).delete()
            db.session.commit()
        
        return jsonify({
            'success': True,
            'order': order.to_dict(),
            'payment_method': 'cash',
            'message': 'Pesanan berhasil dibuat!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def generate_midtrans_snap_token(order, midtrans_order_id):
    """Generate Midtrans Snap token for payment"""
    import requests
    import base64
    
    server_key = app.config.get('MIDTRANS_SERVER_KEY', '')
    is_production = app.config.get('MIDTRANS_IS_PRODUCTION', False)
    
    # Determine API URL
    if is_production:
        snap_url = 'https://app.midtrans.com/snap/v1/transactions'
    else:
        snap_url = 'https://app.sandbox.midtrans.com/snap/v1/transactions'
    
    # Prepare transaction details
    transaction_details = {
        'order_id': midtrans_order_id,
        'gross_amount': int(order.total)
    }
    
    # Prepare item details
    item_details = []
    for item in order.items:
        item_details.append({
            'id': str(item.menu_item_id),
            'price': int(item.price),
            'quantity': item.quantity,
            'name': item.name[:50]  # Midtrans limits name to 50 chars
        })
    
    # Note: Discount is already factored into the order total,
    # no need to add separate line items for it in Midtrans
    
    # Customer details
    customer_details = {
        'first_name': order.customer_name or 'Customer',
        'email': 'customer@kasir.local'
    }
    
    if current_user.is_authenticated:
        customer_details['first_name'] = current_user.full_name or current_user.username
        customer_details['email'] = current_user.email or 'customer@kasir.local'
    
    # Build request payload
    payload = {
        'transaction_details': transaction_details,
        'item_details': item_details,
        'customer_details': customer_details
    }
    
    # Create authorization header
    auth_string = base64.b64encode(f"{server_key}:".encode()).decode()
    headers = {
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'Authorization': f'Basic {auth_string}'
    }
    
    try:
        response = requests.post(snap_url, json=payload, headers=headers, timeout=30)
        if response.status_code == 201:
            data = response.json()
            return data.get('token')
        else:
            print(f"Midtrans error: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Midtrans connection error: {e}")
        return None

@app.route('/api/order/<int:order_id>', methods=['GET'])
@login_required
def api_get_order(order_id):
    """Get order details by ID for the view modal."""
    order = Order.query.get_or_404(order_id)
    return jsonify(order.to_dict())

@app.route('/api/order/<int:order_id>/status', methods=['PUT'])
@login_required
def api_update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.json
    
    new_status = data.get('status', order.status)
    old_status = order.status
    order.status = new_status
    
    # Handle table status
    if new_status in ('completed', 'cancelled') and order.table:
        order.table.status = 'available'
    
    # Restore stock when order is cancelled
    if new_status == 'cancelled' and old_status != 'cancelled':
        for item in order.items:
            if item.menu_item_id:
                menu_item = db.session.get(MenuItem, item.menu_item_id)
                if menu_item:
                    if order.branch_id:
                        bms = get_branch_stock(menu_item.id, order.branch_id)
                        bms.stock += item.quantity
                    else:
                        menu_item.stock += item.quantity
    
    db.session.commit()
    
    return jsonify({'success': True, 'order': order.to_dict()})

@app.route('/api/payment/midtrans', methods=['POST'])
def api_create_midtrans_payment():
    try:
        data = request.json
        order_id = data.get('order_id')
        
        order = Order.query.get_or_404(order_id)
        
        # Create Midtrans transaction (simplified - in production use midtransclient)
        midtrans_order_id = f"DTO-{order.order_number}"
        
        # Update payment
        if order.payment:
            order.payment.payment_method = 'midtrans'
            order.payment.midtrans_order_id = midtrans_order_id
            order.payment.status = 'pending'
            # In production, generate actual Snap token here
            order.payment.payment_url = f"https://app.sandbox.midtrans.com/snap/v2/vtweb/{midtrans_order_id}"
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'payment_url': order.payment.payment_url,
            'order_id': midtrans_order_id
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@csrf.exempt
@app.route('/api/payment/midtrans/callback', methods=['POST'])
@limiter.limit("30 per minute")  # Rate limit webhook calls
def api_midtrans_callback():
    """Handle Midtrans payment notification webhook - CSRF exempt for external service"""
    try:
        data = request.json
        order_id = data.get('order_id')
        transaction_status = data.get('transaction_status')
        
        # Verify signature for security (Midtrans sends signature_key)
        signature_key = data.get('signature_key')
        server_key = app.config.get('MIDTRANS_SERVER_KEY', '')
        
        if signature_key and server_key:
            import hashlib
            # Midtrans signature: SHA512(order_id+status_code+gross_amount+server_key)
            status_code = data.get('status_code', '')
            gross_amount = data.get('gross_amount', '')
            expected_signature = hashlib.sha512(
                f"{order_id}{status_code}{gross_amount}{server_key}".encode()
            ).hexdigest()
            
            if signature_key != expected_signature:
                return jsonify({'error': 'Invalid signature'}), 403
        
        # Find payment by midtrans_order_id
        payment = Payment.query.filter_by(midtrans_order_id=order_id).first()
        
        if payment:
            payment.midtrans_status = transaction_status
            
            if transaction_status in ['capture', 'settlement']:
                payment.status = 'paid'
                payment.paid_at = utc_now()
                payment.order.status = 'processing'
                
                # Create success notification
                amount_formatted = f"{payment.amount:,}".replace(',', '.')
                create_notification(
                    type='payment_success',
                    title='Pembayaran Berhasil!',
                    message=f'Order #{payment.order.order_number} - Rp {amount_formatted}',
                    data={'order_id': payment.order_id, 'payment_id': payment.id}
                )
                
            elif transaction_status in ['deny', 'cancel', 'expire']:
                payment.status = 'failed'
                
                # Create failure notification
                create_notification(
                    type='payment_failed',
                    title='Pembayaran Gagal',
                    message=f'Order #{payment.order.order_number} - Status: {transaction_status}',
                    data={'order_id': payment.order_id, 'payment_id': payment.id}
                )
            
            db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Statistics API
@app.route('/api/stats')
@login_required
def api_get_stats():
    today = datetime.now().date()
    
    # Today's stats
    today_orders = branch_filter(Order.query, Order).filter(
        db.func.date(Order.created_at) == today
    ).all()
    
    total_income = sum(o.total for o in today_orders if o.payment and o.payment.status == 'paid')
    total_orders = len(today_orders)
    
    # Popular items today
    popular_items = {}
    for order in today_orders:
        for item in order.items:
            if item.name not in popular_items:
                popular_items[item.name] = 0
            popular_items[item.name] += item.quantity
    
    most_popular = sorted(popular_items.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return jsonify({
        'total_income': total_income,
        'total_orders': total_orders,
        'most_popular': most_popular,
        'average_transaction': total_income / total_orders if total_orders > 0 else 0
    })

# Profile
@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    full_name = request.form.get('full_name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    
    # Check email uniqueness
    if email != current_user.email:
        if User.query.filter_by(email=email).first():
            flash('Email sudah digunakan!', 'danger')
            return redirect(url_for('profile'))
    
    current_user.full_name = full_name
    current_user.email = email
    current_user.phone = phone
    
    db.session.commit()
    flash('Profil berhasil diperbarui!', 'success')
    return redirect(url_for('profile'))

@app.route('/profile/change-password', methods=['POST'])
@login_required
def profile_change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not current_user.check_password(current_password):
        flash('Password saat ini salah!', 'danger')
        return redirect(url_for('profile'))
    
    if new_password != confirm_password:
        flash('Password baru tidak cocok!', 'danger')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('Password minimal 6 karakter!', 'danger')
        return redirect(url_for('profile'))
    
    current_user.set_password(new_password)
    db.session.commit()
    flash('Password berhasil diubah!', 'success')
    return redirect(url_for('profile'))

# Admin routes
@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    users = branch_filter(User.query, User).all()
    roles = Role.query.all()
    return render_template('admin/users.html', users=users, roles=roles)

@app.route('/admin/users/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    role_id = request.form.get('role_id')
    
    if User.query.filter_by(username=username).first():
        flash('Username sudah digunakan!', 'danger')
        return redirect(url_for('admin_users'))
    
    user = User(
        username=username,
        email=email,
        full_name=full_name,
        branch_id=int(request.form['branch_id']) if request.form.get('branch_id', '').strip() else get_user_branch_id()
    )
    user.set_password(password)
    
    if role_id:
        role = db.session.get(Role, role_id)
        if role:
            user.roles.append(role)
    
    db.session.add(user)
    db.session.commit()
    
    flash('User berhasil dibuat!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'diaktifkan' if user.is_active else 'dinonaktifkan'
    flash(f'User {user.username} berhasil {status}!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/menu')
@login_required
@role_required('admin', 'manager')
def admin_menu():
    categories = Category.query.order_by(Category.order).all()
    menu_items = MenuItem.query.all()
    return render_template('admin/menu.html', categories=categories, menu_items=menu_items)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'gif', 'webp'})

def save_uploaded_image(file):
    """Save uploaded image and return the URL path"""
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # Add timestamp to avoid duplicate names
        import time
        filename = f"{int(time.time())}_{filename}"
        
        # Ensure upload folder exists
        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
        menu_upload_folder = os.path.join(upload_folder, 'menu')
        os.makedirs(menu_upload_folder, exist_ok=True)
        
        filepath = os.path.join(menu_upload_folder, filename)
        file.save(filepath)
        
        # Return URL path
        return f"/uploads/menu/{filename}"
    return None

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve uploaded files"""
    upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
    return send_file(os.path.join(upload_folder, filename))

@app.route('/admin/menu/create', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_create_menu():
    code = request.form.get('code')
    name = request.form.get('name')
    price = int(request.form.get('price', 0))
    category_id = request.form.get('category_id')
    description = request.form.get('description')
    is_popular = request.form.get('is_popular') == 'on'
    has_spicy_option = request.form.get('has_spicy_option') == 'on'
    has_temperature_option = request.form.get('has_temperature_option') == 'on'
    
    # Handle image: URL or upload
    image = request.form.get('image_url', '').strip()
    if 'image_file' in request.files:
        file = request.files['image_file']
        if file and file.filename:
            uploaded_path = save_uploaded_image(file)
            if uploaded_path:
                image = uploaded_path
    
    # Default image if none provided
    if not image:
        image = "https://via.placeholder.com/300x200?text=No+Image"
    
    menu_item = MenuItem(
        code=code,
        name=name,
        price=price,
        category_id=category_id,
        description=description,
        is_popular=is_popular,
        has_spicy_option=has_spicy_option,
        has_temperature_option=has_temperature_option,
        image=image
    )
    
    db.session.add(menu_item)
    db.session.flush()
    
    # Create BranchMenuStock entries for all branches (including inactive)
    branches = Branch.query.all()
    for branch in branches:
        bms = BranchMenuStock(branch_id=branch.id, menu_item_id=menu_item.id, stock=100, is_available=True)
        db.session.add(bms)
    
    db.session.commit()
    
    flash('Menu berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_menu'))


@app.route('/admin/menu/<int:id>/edit', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_edit_menu(id):
    menu_item = MenuItem.query.get_or_404(id)
    
    menu_item.code = request.form.get('code', menu_item.code)
    menu_item.name = request.form.get('name', menu_item.name)
    menu_item.price = int(request.form.get('price', menu_item.price))
    menu_item.category_id = request.form.get('category_id', menu_item.category_id)
    menu_item.description = request.form.get('description', menu_item.description)
    menu_item.is_popular = request.form.get('is_popular') == 'on'
    menu_item.has_spicy_option = request.form.get('has_spicy_option') == 'on'
    menu_item.has_temperature_option = request.form.get('has_temperature_option') == 'on'
    
    # Update per-branch stock/availability
    bid = get_user_branch_id()
    new_available = request.form.get('is_available') == 'on'
    if bid:
        bms = get_branch_stock(menu_item.id, bid)
        bms.is_available = new_available
    else:
        # Owner: update global is_available + all branches
        menu_item.is_available = new_available
        BranchMenuStock.query.filter_by(menu_item_id=menu_item.id).update({'is_available': new_available})
    
    # Handle image: URL or upload
    image_url = request.form.get('image_url', '').strip()
    if image_url:
        menu_item.image = image_url
    
    if 'image_file' in request.files:
        file = request.files['image_file']
        if file and file.filename:
            uploaded_path = save_uploaded_image(file)
            if uploaded_path:
                menu_item.image = uploaded_path
    
    db.session.commit()
    
    flash('Menu berhasil diperbarui!', 'success')
    return redirect(url_for('admin_menu'))


@app.route('/admin/menu/<int:id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_delete_menu(id):
    menu_item = MenuItem.query.get_or_404(id)
    
    # Check if menu item is used in any orders
    order_items = OrderItem.query.filter_by(menu_item_id=id).first()
    if order_items:
        flash('Menu tidak dapat dihapus karena sudah digunakan dalam pesanan. Nonaktifkan saja jika tidak ingin ditampilkan.', 'error')
        return redirect(url_for('admin_menu'))
    
    # Delete from cart items first
    CartItem.query.filter_by(menu_item_id=id).delete()
    
    # Delete branch stock entries
    BranchMenuStock.query.filter_by(menu_item_id=id).delete()
    
    db.session.delete(menu_item)
    db.session.commit()
    
    flash('Menu berhasil dihapus!', 'success')
    return redirect(url_for('admin_menu'))


@app.route('/api/menu/<int:id>')
@login_required
@role_required('admin', 'manager')
def api_get_menu_item(id):
    """Get menu item data for edit form"""
    menu_item = MenuItem.query.get_or_404(id)
    bid = get_user_branch_id()
    data = {
        'id': menu_item.id,
        'code': menu_item.code,
        'name': menu_item.name,
        'price': menu_item.price,
        'category_id': menu_item.category_id,
        'description': menu_item.description or '',
        'image': menu_item.image or '',
        'is_popular': menu_item.is_popular,
        'is_available': menu_item.is_available,
        'has_spicy_option': menu_item.has_spicy_option,
        'has_temperature_option': menu_item.has_temperature_option
    }
    # Return per-branch availability if user is branch-specific
    if bid:
        bms = get_branch_stock(menu_item.id, bid)
        data['is_available'] = bms.is_available
        data['stock'] = bms.stock
    return jsonify(data)

@app.route('/api/menu/<int:id>/toggle', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def api_toggle_menu(id):
    """Toggle menu item availability (on/off / habis)"""
    menu_item = MenuItem.query.get_or_404(id)
    bid = get_user_branch_id()
    
    if bid:
        # Branch user: toggle per-branch availability
        bms = get_branch_stock(menu_item.id, bid)
        bms.is_available = not bms.is_available
        new_status = bms.is_available
    else:
        # Owner: toggle global + all branches
        menu_item.is_available = not menu_item.is_available
        new_status = menu_item.is_available
        BranchMenuStock.query.filter_by(menu_item_id=menu_item.id).update({'is_available': new_status})
    
    db.session.commit()
    return jsonify({'success': True, 'is_available': new_status, 'name': menu_item.name})

@app.route('/admin/printer')
@login_required
@role_required('admin', 'manager')
def admin_printer():
    """Redirect to Printer Station - the dedicated printer page"""
    return redirect(url_for('printer_station'))


@app.route('/printer-station')
@login_required
def printer_station():
    """Dedicated printer station page - keep this open for reliable printing"""
    return render_template('printer_station.html')


@app.route('/admin/tables')
@login_required
@role_required('admin', 'manager')
def admin_tables():
    tables = Table.query.all()
    return render_template('admin/tables.html', tables=tables)

@app.route('/admin/tables/<int:table_id>/qr')
@login_required
@role_required('admin', 'manager')
def admin_table_qr(table_id):
    table = Table.query.get_or_404(table_id)
    qr_path, qr_base64 = generate_table_qr(table.number)
    table.qr_code = qr_path
    db.session.commit()
    
    return jsonify({
        'success': True,
        'qr_code': f"data:image/png;base64,{qr_base64}",
        'table_number': table.number
    })

@app.route('/admin/tables/add', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_table_add():
    """Add new table"""
    number = request.form.get('number', '').strip()
    name = request.form.get('name', '').strip()
    capacity = int(request.form.get('capacity', 4))
    
    if not number:
        flash('Nomor meja wajib diisi!', 'danger')
        return redirect(url_for('admin_tables'))
    
    if Table.query.filter_by(number=number).first():
        flash('Nomor meja sudah ada!', 'danger')
        return redirect(url_for('admin_tables'))
    
    table = Table(
        number=number,
        name=name or f"Meja {number}",
        capacity=capacity
    )
    db.session.add(table)
    db.session.commit()
    
    flash(f'Meja {number} berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_tables'))

@app.route('/admin/tables/<int:table_id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_table_delete(table_id):
    """Delete a table"""
    table = Table.query.get_or_404(table_id)
    
    # Check if table has active orders
    has_active_order = Order.query.filter_by(table_id=table_id).filter(
        Order.status.in_(['pending', 'processing'])
    ).first() is not None
    
    if has_active_order:
        flash('Tidak bisa hapus meja dengan pesanan aktif!', 'danger')
        return redirect(url_for('admin_tables'))
    
    table_num = table.number
    db.session.delete(table)
    db.session.commit()
    
    flash(f'Meja {table_num} berhasil dihapus!', 'success')
    return redirect(url_for('admin_tables'))

@app.route('/admin/tables/<int:table_id>/toggle', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_table_toggle(table_id):
    """Toggle table status between available and occupied"""
    table = Table.query.get_or_404(table_id)
    
    if table.status == 'available':
        table.status = 'occupied'
    else:
        table.status = 'available'
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'status': table.status,
        'message': f'Meja {table.number} status: {table.status}'
    })


# ========== DISCOUNT/PROMO MANAGEMENT ==========
@app.route('/admin/discounts')
@login_required
@role_required('admin', 'manager')
def admin_discounts():
    """Discount management page"""
    discounts = branch_filter(Discount.query, Discount).order_by(Discount.created_at.desc()).all()
    return render_template('admin/discounts.html', discounts=discounts)


@app.route('/admin/discounts/create', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_discount_create():
    """Create new discount/promo"""
    try:
        name = request.form.get('name')
        code = request.form.get('code', '').upper().strip()
        description = request.form.get('description', '')
        discount_type = request.form.get('discount_type', 'percentage')
        value = int(request.form.get('value', 0))
        min_purchase = int(request.form.get('min_purchase', 0))
        max_discount = request.form.get('max_discount')
        usage_limit = request.form.get('usage_limit')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        is_active = request.form.get('is_active') == 'on'
        
        # Validate required fields
        if not name or not code or value <= 0:
            flash('Nama, kode, dan nilai diskon harus diisi dengan benar', 'danger')
            return redirect(url_for('admin_discounts'))
        
        # Check if code already exists
        existing = Discount.query.filter_by(code=code).first()
        if existing:
            flash(f'Kode promo "{code}" sudah digunakan', 'danger')
            return redirect(url_for('admin_discounts'))
        
        # Parse optional fields
        max_discount = int(max_discount) if max_discount else None
        usage_limit = int(usage_limit) if usage_limit else None
        start_date = datetime.fromisoformat(start_date) if start_date else None
        end_date = datetime.fromisoformat(end_date) if end_date else None
        
        discount = Discount(
            name=name,
            code=code,
            description=description,
            discount_type=discount_type,
            value=value,
            min_purchase=min_purchase,
            max_discount=max_discount,
            usage_limit=usage_limit,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active,
            branch_id=get_default_branch_id()
        )
        
        db.session.add(discount)
        db.session.commit()
        
        flash(f'Promo "{name}" berhasil ditambahkan', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menambahkan promo: {str(e)}', 'danger')
    
    return redirect(url_for('admin_discounts'))


@app.route('/admin/discounts/<int:discount_id>/edit', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_discount_edit(discount_id):
    """Edit existing discount"""
    discount = Discount.query.get_or_404(discount_id)
    
    try:
        discount.name = request.form.get('name', discount.name)
        discount.description = request.form.get('description', '')
        discount.discount_type = request.form.get('discount_type', discount.discount_type)
        discount.value = int(request.form.get('value', discount.value))
        discount.min_purchase = int(request.form.get('min_purchase', 0))
        
        max_discount = request.form.get('max_discount')
        discount.max_discount = int(max_discount) if max_discount else None
        
        usage_limit = request.form.get('usage_limit')
        discount.usage_limit = int(usage_limit) if usage_limit else None
        
        start_date = request.form.get('start_date')
        discount.start_date = datetime.fromisoformat(start_date) if start_date else None
        
        end_date = request.form.get('end_date')
        discount.end_date = datetime.fromisoformat(end_date) if end_date else None
        
        discount.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        flash(f'Promo "{discount.name}" berhasil diperbarui', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal memperbarui promo: {str(e)}', 'danger')
    
    return redirect(url_for('admin_discounts'))


@app.route('/admin/discounts/<int:discount_id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_discount_delete(discount_id):
    """Delete a discount"""
    discount = Discount.query.get_or_404(discount_id)
    
    try:
        name = discount.name
        db.session.delete(discount)
        db.session.commit()
        flash(f'Promo "{name}" berhasil dihapus', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus promo: {str(e)}', 'danger')
    
    return redirect(url_for('admin_discounts'))


@app.route('/admin/discounts/<int:discount_id>/toggle', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_discount_toggle(discount_id):
    """Toggle discount active status"""
    discount = Discount.query.get_or_404(discount_id)
    discount.is_active = not discount.is_active
    db.session.commit()
    
    status = "aktif" if discount.is_active else "nonaktif"
    return jsonify({
        'success': True,
        'is_active': discount.is_active,
        'message': f'Promo "{discount.name}" sekarang {status}'
    })


@app.route('/api/discount/validate', methods=['POST'])
@login_required
def api_validate_discount():
    """Validate a discount code for given subtotal"""
    data = request.get_json()
    code = data.get('code', '').upper().strip()
    subtotal = int(data.get('subtotal', 0))
    
    if not code:
        return jsonify({'valid': False, 'message': 'Kode promo tidak boleh kosong'})
    
    discount = Discount.query.filter_by(code=code).first()
    if not discount:
        return jsonify({'valid': False, 'message': 'Kode promo tidak ditemukan'})
    
    is_valid, message = discount.is_valid(subtotal)
    if not is_valid:
        return jsonify({'valid': False, 'message': message})
    
    discount_amount = discount.calculate_discount(subtotal)
    
    return jsonify({
        'valid': True,
        'message': 'Promo berhasil digunakan!',
        'discount': discount.to_dict(),
        'discount_amount': discount_amount
    })


@app.route('/api/discounts/active')
@login_required
def api_active_discounts():
    """Get all currently active discounts"""
    now = utc_now()
    discounts = branch_filter(Discount.query, Discount).filter(
        Discount.is_active == True,
        (Discount.start_date == None) | (Discount.start_date <= now),
        (Discount.end_date == None) | (Discount.end_date >= now),
        (Discount.usage_limit == None) | (Discount.usage_count < Discount.usage_limit)
    ).all()
    
    return jsonify({
        'discounts': [d.to_dict() for d in discounts]
    })


# ============================================
# BRANCH / CABANG MANAGEMENT
# ============================================

@app.route('/admin/branches')
@login_required
@role_required('admin', 'manager')
def admin_branches():
    """Branch management page"""
    branches = Branch.query.order_by(Branch.created_at.desc()).all()
    cities = City.query.order_by(City.name).all()
    brands = Brand.query.order_by(Brand.name).all()
    return render_template('admin/branches.html', branches=branches, cities=cities, brands=brands, active_page='admin_branches')

@app.route('/admin/branches/create', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_branch_create():
    """Create a new branch"""
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    address = request.form.get('address', '').strip()
    phone = request.form.get('phone', '').strip()
    manager_name = request.form.get('manager_name', '').strip()
    opening_time = request.form.get('opening_time', '08:00').strip()
    closing_time = request.form.get('closing_time', '22:00').strip()
    city_id = request.form.get('city_id', '').strip()
    brand_id = request.form.get('brand_id', '').strip()
    
    if not name or not code:
        flash('Nama dan kode cabang wajib diisi.', 'danger')
        return redirect(url_for('admin_branches'))
    
    # Check for duplicate code
    existing = Branch.query.filter_by(code=code).first()
    if existing:
        flash('Kode cabang sudah digunakan.', 'danger')
        return redirect(url_for('admin_branches'))
    
    branch = Branch(
        name=name,
        code=code,
        address=address,
        phone=phone,
        manager_name=manager_name,
        opening_time=opening_time,
        closing_time=closing_time,
        city_id=int(city_id) if city_id and city_id.isdigit() else None,
        brand_id=int(brand_id) if brand_id and brand_id.isdigit() else None
    )
    db.session.add(branch)
    db.session.flush()
    
    # Create BranchMenuStock entries for all existing menu items
    all_menu_items = MenuItem.query.all()
    for mi in all_menu_items:
        bms = BranchMenuStock(branch_id=branch.id, menu_item_id=mi.id, stock=100, is_available=True)
        db.session.add(bms)
    
    db.session.commit()
    
    flash(f'Cabang "{name}" berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_branches'))

@app.route('/admin/branches/<int:branch_id>/edit', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_branch_edit(branch_id):
    """Edit a branch"""
    branch = db.session.get(Branch, branch_id)
    if not branch:
        flash('Cabang tidak ditemukan.', 'danger')
        return redirect(url_for('admin_branches'))
    
    branch.name = request.form.get('name', branch.name).strip()
    new_code = request.form.get('code', branch.code).strip().upper()
    
    # Check for duplicate code (exclude current branch)
    existing = Branch.query.filter(Branch.code == new_code, Branch.id != branch_id).first()
    if existing:
        flash('Kode cabang sudah digunakan.', 'danger')
        return redirect(url_for('admin_branches'))
    
    branch.code = new_code
    branch.address = request.form.get('address', branch.address).strip()
    branch.phone = request.form.get('phone', branch.phone).strip()
    branch.manager_name = request.form.get('manager_name', branch.manager_name).strip()
    branch.opening_time = request.form.get('opening_time', branch.opening_time).strip()
    branch.closing_time = request.form.get('closing_time', branch.closing_time).strip()
    city_id_str = request.form.get('city_id', '').strip()
    brand_id_str = request.form.get('brand_id', '').strip()
    branch.city_id = int(city_id_str) if city_id_str and city_id_str.isdigit() else None
    branch.brand_id = int(brand_id_str) if brand_id_str and brand_id_str.isdigit() else None
    
    db.session.commit()
    flash(f'Cabang "{branch.name}" berhasil diperbarui!', 'success')
    return redirect(url_for('admin_branches'))

@app.route('/admin/branches/<int:branch_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_branch_toggle(branch_id):
    """Toggle branch active status - only owner (admin with no branch) can do this"""
    # Only owner/admin pusat (branch_id=NULL) can toggle branches
    if current_user.branch_id is not None:
        return jsonify({'success': False, 'error': 'Hanya owner/admin pusat yang dapat mengubah status cabang'}), 403
    
    branch = db.session.get(Branch, branch_id)
    if not branch:
        return jsonify({'success': False, 'error': 'Cabang tidak ditemukan'}), 404
    
    # Prevent deactivating the Pusat branch
    if branch.code == 'PUSAT' and branch.is_active:
        return jsonify({'success': False, 'error': 'Cabang Pusat tidak dapat dinonaktifkan'}), 400
    
    branch.is_active = not branch.is_active
    db.session.commit()
    
    status = 'aktif' if branch.is_active else 'nonaktif'
    return jsonify({
        'success': True,
        'is_active': branch.is_active,
        'message': f'Cabang "{branch.name}" sekarang {status}'
    })

@app.route('/admin/branches/<int:branch_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_branch_delete(branch_id):
    """Delete a branch - only owner (admin with no branch) can do this"""
    # Only owner/admin pusat (branch_id=NULL) can delete branches
    if current_user.branch_id is not None:
        flash('Hanya owner/admin pusat yang dapat menghapus cabang.', 'danger')
        return redirect(url_for('admin_branches'))
    
    branch = db.session.get(Branch, branch_id)
    if not branch:
        flash('Cabang tidak ditemukan.', 'danger')
        return redirect(url_for('admin_branches'))
    
    # Prevent deleting the Pusat branch
    if branch.code == 'PUSAT':
        flash('Cabang Pusat tidak dapat dihapus.', 'danger')
        return redirect(url_for('admin_branches'))
    
    name = branch.name
    # Clean up BranchMenuStock entries first (NOT NULL constraint on branch_id)
    BranchMenuStock.query.filter_by(branch_id=branch_id).delete()
    db.session.delete(branch)
    db.session.commit()
    
    flash(f'Cabang "{name}" berhasil dihapus.', 'success')
    return redirect(url_for('admin_branches'))


# ============================================
# CITY & BRAND MANAGEMENT (MULTI-OUTLET)
# ============================================

@app.route('/admin/cities')
@login_required
@role_required('admin')
def admin_cities():
    """City management page"""
    cities = City.query.order_by(City.name).all()
    return render_template('admin/cities.html', cities=cities, active_page='admin_cities')

@app.route('/admin/cities/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_city_create():
    """Create a new city"""
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    
    if not name or not code:
        flash('Nama dan kode kota wajib diisi.', 'danger')
        return redirect(url_for('admin_cities'))
    
    if City.query.filter_by(code=code).first():
        flash('Kode kota sudah digunakan.', 'danger')
        return redirect(url_for('admin_cities'))
    
    city = City(name=name, code=code)
    db.session.add(city)
    db.session.commit()
    
    flash(f'Kota "{name}" berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_cities'))

@app.route('/admin/cities/<int:city_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def admin_city_edit(city_id):
    """Edit a city"""
    city = db.session.get(City, city_id)
    if not city:
        flash('Kota tidak ditemukan.', 'danger')
        return redirect(url_for('admin_cities'))
    
    city.name = request.form.get('name', city.name).strip()
    new_code = request.form.get('code', city.code).strip().upper()
    
    existing = City.query.filter(City.code == new_code, City.id != city_id).first()
    if existing:
        flash('Kode kota sudah digunakan.', 'danger')
        return redirect(url_for('admin_cities'))
    
    city.code = new_code
    db.session.commit()
    flash(f'Kota "{city.name}" berhasil diperbarui!', 'success')
    return redirect(url_for('admin_cities'))

@app.route('/admin/cities/<int:city_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_city_toggle(city_id):
    """Toggle city active status"""
    if current_user.branch_id is not None:
        return jsonify({'success': False, 'error': 'Hanya owner yang dapat mengubah status kota'}), 403
    
    city = db.session.get(City, city_id)
    if not city:
        return jsonify({'success': False, 'error': 'Kota tidak ditemukan'}), 404
    
    city.is_active = not city.is_active
    db.session.commit()
    
    return jsonify({
        'success': True,
        'is_active': city.is_active,
        'message': f'Kota "{city.name}" sekarang {"aktif" if city.is_active else "nonaktif"}'
    })

@app.route('/admin/cities/<int:city_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_city_delete(city_id):
    """Delete a city"""
    if current_user.branch_id is not None:
        flash('Hanya owner yang dapat menghapus kota.', 'danger')
        return redirect(url_for('admin_cities'))
    
    city = db.session.get(City, city_id)
    if not city:
        flash('Kota tidak ditemukan.', 'danger')
        return redirect(url_for('admin_cities'))
    
    if city.branches.count() > 0:
        flash(f'Kota "{city.name}" masih memiliki {city.branches.count()} cabang. Hapus atau pindahkan cabang terlebih dahulu.', 'danger')
        return redirect(url_for('admin_cities'))
    
    name = city.name
    db.session.delete(city)
    db.session.commit()
    flash(f'Kota "{name}" berhasil dihapus.', 'success')
    return redirect(url_for('admin_cities'))

@app.route('/admin/brands')
@login_required
@role_required('admin')
def admin_brands():
    """Brand management page"""
    brands = Brand.query.order_by(Brand.name).all()
    return render_template('admin/brands.html', brands=brands, active_page='admin_brands')

@app.route('/admin/brands/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_brand_create():
    """Create a new brand"""
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    description = request.form.get('description', '').strip()
    
    if not name or not code:
        flash('Nama dan kode brand wajib diisi.', 'danger')
        return redirect(url_for('admin_brands'))
    
    if Brand.query.filter_by(code=code).first():
        flash('Kode brand sudah digunakan.', 'danger')
        return redirect(url_for('admin_brands'))
    
    brand = Brand(name=name, code=code, description=description)
    db.session.add(brand)
    db.session.commit()
    
    flash(f'Brand "{name}" berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_brands'))

@app.route('/admin/brands/<int:brand_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def admin_brand_edit(brand_id):
    """Edit a brand"""
    brand = db.session.get(Brand, brand_id)
    if not brand:
        flash('Brand tidak ditemukan.', 'danger')
        return redirect(url_for('admin_brands'))
    
    brand.name = request.form.get('name', brand.name).strip()
    new_code = request.form.get('code', brand.code).strip().upper()
    
    existing = Brand.query.filter(Brand.code == new_code, Brand.id != brand_id).first()
    if existing:
        flash('Kode brand sudah digunakan.', 'danger')
        return redirect(url_for('admin_brands'))
    
    brand.code = new_code
    brand.description = request.form.get('description', brand.description or '').strip()
    db.session.commit()
    flash(f'Brand "{brand.name}" berhasil diperbarui!', 'success')
    return redirect(url_for('admin_brands'))

@app.route('/admin/brands/<int:brand_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_brand_toggle(brand_id):
    """Toggle brand active status"""
    if current_user.branch_id is not None:
        return jsonify({'success': False, 'error': 'Hanya owner yang dapat mengubah status brand'}), 403
    
    brand = db.session.get(Brand, brand_id)
    if not brand:
        return jsonify({'success': False, 'error': 'Brand tidak ditemukan'}), 404
    
    brand.is_active = not brand.is_active
    db.session.commit()
    
    return jsonify({
        'success': True,
        'is_active': brand.is_active,
        'message': f'Brand "{brand.name}" sekarang {"aktif" if brand.is_active else "nonaktif"}'
    })

@app.route('/admin/brands/<int:brand_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_brand_delete(brand_id):
    """Delete a brand"""
    if current_user.branch_id is not None:
        flash('Hanya owner yang dapat menghapus brand.', 'danger')
        return redirect(url_for('admin_brands'))
    
    brand = db.session.get(Brand, brand_id)
    if not brand:
        flash('Brand tidak ditemukan.', 'danger')
        return redirect(url_for('admin_brands'))
    
    if brand.branches.count() > 0:
        flash(f'Brand "{brand.name}" masih memiliki {brand.branches.count()} cabang. Hapus atau pindahkan cabang terlebih dahulu.', 'danger')
        return redirect(url_for('admin_brands'))
    
    name = brand.name
    db.session.delete(brand)
    db.session.commit()
    flash(f'Brand "{name}" berhasil dihapus.', 'success')
    return redirect(url_for('admin_brands'))


# ============================================
# EXPENSE TRACKING (PENGELUARAN)
# ============================================

@app.route('/expenses')
@login_required
@role_required('admin', 'manager', 'kasir')
def expenses():
    """Expense tracking page"""
    from datetime import date
    
    # Get filter parameters
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    category_filter = request.args.get('category', 'all')
    
    try:
        filter_year, filter_month = map(int, month.split('-'))
    except (ValueError, AttributeError):
        filter_year, filter_month = date.today().year, date.today().month
    
    # Build query
    query = branch_filter(Expense.query, Expense).filter(
        db.extract('year', Expense.date) == filter_year,
        db.extract('month', Expense.date) == filter_month
    )
    
    if category_filter != 'all':
        query = query.filter(Expense.category == category_filter)
    
    expense_list = query.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    
    # Calculate totals
    total_expenses = sum(e.amount for e in expense_list)
    category_totals = {}
    for e in expense_list:
        label = Expense.CATEGORIES.get(e.category, e.category)
        category_totals[label] = category_totals.get(label, 0) + e.amount
    
    return render_template('expenses.html',
                         expenses=expense_list,
                         total_expenses=total_expenses,
                         category_totals=category_totals,
                         expense_categories=Expense.CATEGORIES,
                         current_month=month,
                         category_filter=category_filter,
                         active_page='expenses',
                         now=datetime.now())

@app.route('/expenses/add', methods=['POST'])
@login_required
@role_required('admin', 'manager', 'kasir')
def expense_add():
    """Add a new expense"""
    from datetime import date as date_type
    
    expense_date = request.form.get('date')
    category = request.form.get('category', '').strip()
    description = request.form.get('description', '').strip()
    amount = request.form.get('amount', '0')
    notes = request.form.get('notes', '').strip()
    
    if not description or not category:
        flash('Deskripsi dan kategori wajib diisi.', 'danger')
        return redirect(url_for('expenses'))
    
    try:
        amount = int(amount)
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('Jumlah harus angka positif.', 'danger')
        return redirect(url_for('expenses'))
    
    try:
        parsed_date = date_type.fromisoformat(expense_date) if expense_date else date_type.today()
    except ValueError:
        parsed_date = date_type.today()
    
    expense = Expense(
        date=parsed_date,
        category=category,
        description=description,
        amount=amount,
        notes=notes,
        user_id=current_user.id,
        branch_id=get_default_branch_id()
    )
    db.session.add(expense)
    db.session.commit()
    
    flash(f'Pengeluaran "{description}" berhasil ditambahkan!', 'success')
    return redirect(url_for('expenses'))

@app.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def expense_delete(expense_id):
    """Delete an expense"""
    expense = db.session.get(Expense, expense_id)
    if not expense:
        flash('Pengeluaran tidak ditemukan.', 'danger')
        return redirect(url_for('expenses'))
    
    db.session.delete(expense)
    db.session.commit()
    
    flash('Pengeluaran berhasil dihapus.', 'success')
    return redirect(url_for('expenses'))


# ============================================
# SHIFT MANAGEMENT
# ============================================

@app.route('/api/shift/open', methods=['POST'])
@login_required
@role_required('admin', 'manager', 'kasir')
def api_shift_open():
    """Open a new cashier shift"""
    # Check if user already has an open shift
    open_shift = CashierShift.query.filter_by(user_id=current_user.id, status='open').first()
    if open_shift:
        return jsonify({'success': False, 'error': 'Anda sudah memiliki shift yang aktif'}), 400
    
    data = request.get_json()
    opening_cash = int(data.get('opening_cash', 0))
    
    shift = CashierShift(
        user_id=current_user.id,
        opening_cash=opening_cash,
        notes=data.get('notes', ''),
        branch_id=get_default_branch_id()
    )
    db.session.add(shift)
    db.session.commit()
    
    return jsonify({'success': True, 'shift': shift.to_dict()})

@app.route('/api/shift/close', methods=['POST'])
@login_required
@role_required('admin', 'manager', 'kasir')
def api_shift_close():
    """Close the current cashier shift"""
    shift = CashierShift.query.filter_by(user_id=current_user.id, status='open').first()
    if not shift:
        return jsonify({'success': False, 'error': 'Tidak ada shift yang aktif'}), 400
    
    data = request.get_json()
    shift.closing_cash = int(data.get('closing_cash', 0))
    shift.end_time = utc_now()
    shift.status = 'closed'
    shift.notes = data.get('notes', shift.notes)
    
    # Calculate total sales during shift using database aggregation
    from sqlalchemy import func
    shift_stats = db.session.query(
        func.count(Order.id).label('total_orders'),
        func.coalesce(func.sum(Order.total), 0).label('total_sales')
    ).filter(
        Order.created_at >= shift.start_time,
        Order.created_at <= shift.end_time,
        Order.user_id == current_user.id,
        Order.status.in_(['processing', 'completed'])
    ).first()
    
    shift.total_orders = shift_stats.total_orders
    shift.total_sales = shift_stats.total_sales
    
    db.session.commit()
    
    return jsonify({'success': True, 'shift': shift.to_dict()})

@app.route('/api/shift/current')
@login_required
def api_shift_current():
    """Get current open shift for the user"""
    shift = CashierShift.query.filter_by(user_id=current_user.id, status='open').first()
    return jsonify({
        'success': True,
        'shift': shift.to_dict() if shift else None
    })
@app.route('/kitchen')
@login_required
@role_required('admin', 'manager', 'koki')
def kitchen():
    """Kitchen display page for cooks"""
    return render_template('kitchen.html')

@app.route('/api/kitchen/orders')
@login_required
@role_required('admin', 'manager', 'koki', 'kasir')
def api_kitchen_orders():
    """Get orders for kitchen display"""
    # Get orders from today that are not completed/cancelled
    today = utc_now().date()
    orders = branch_filter(Order.query, Order).filter(
        Order.created_at >= datetime.combine(today, datetime.min.time()),
        Order.status.in_(['pending', 'processing'])
    ).order_by(Order.created_at.asc()).all()
    
    result = []
    for order in orders:
        order_data = {
            'id': order.id,
            'order_number': order.order_number,
            'table': order.table.number if order.table else 'Takeaway',
            'customer_name': order.customer_name or 'Guest',
            'order_type': order.order_type,
            'status': order.status,
            'created_at': order.created_at.strftime('%H:%M'),
            'items': []
        }
        
        for item in order.items:
            order_data['items'].append({
                'id': item.id,
                'name': item.name,
                'quantity': item.quantity,
                'spice_level': item.spice_level,
                'temperature': item.temperature,
                'notes': item.notes,
                'item_status': item.item_status or 'pending'
            })
        
        result.append(order_data)
    
    return jsonify(result)

@app.route('/api/kitchen/item/<int:item_id>/status', methods=['PUT'])
@login_required
@role_required('admin', 'manager', 'koki', 'kasir')
def api_update_item_status(item_id):
    """Update individual item status"""
    item = OrderItem.query.get_or_404(item_id)
    data = request.get_json()
    new_status = data.get('status')
    
    if new_status not in ['pending', 'cooking', 'ready', 'served']:
        return jsonify({'error': 'Invalid status'}), 400
    
    item.item_status = new_status
    db.session.commit()
    
    # Check if all items in order are ready/served, update order status
    order = item.order
    all_items_status = [i.item_status for i in order.items]
    
    if all(s == 'served' for s in all_items_status):
        order.status = 'completed'
        if order.table:
            order.table.status = 'available'
    elif all(s in ['ready', 'served'] for s in all_items_status):
        order.status = 'processing'  # Ready to serve
    elif any(s == 'cooking' for s in all_items_status):
        order.status = 'processing'
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'item_id': item_id,
        'status': new_status,
        'order_status': order.status
    })

@app.route('/api/kitchen/order/<int:order_id>/status', methods=['PUT'])
@login_required
@role_required('admin', 'manager', 'koki', 'kasir')
def api_update_order_kitchen_status(order_id):
    """Update all items in an order to a status"""
    order = Order.query.get_or_404(order_id)
    data = request.get_json()
    new_status = data.get('status')
    
    if new_status not in ['pending', 'cooking', 'ready', 'served']:
        return jsonify({'error': 'Invalid status'}), 400
    
    # Update all items
    for item in order.items:
        item.item_status = new_status
    
    # Update order status
    if new_status == 'served':
        order.status = 'completed'
        if order.table:
            order.table.status = 'available'
    elif new_status in ['cooking', 'ready']:
        order.status = 'processing'
    else:
        order.status = 'pending'
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'order_id': order_id,
        'item_status': new_status,
        'order_status': order.status
    })

# Printer Status API
@app.route('/api/printer-status')
@login_required
def get_printer_status():
    """Get saved printer info for current user"""
    return jsonify({
        'success': True,
        'printer_name': current_user.printer_name,
        'printer_id': current_user.printer_id
    })

@app.route('/api/printer-status', methods=['POST'])
@login_required
def save_printer_status():
    """Save printer info to database for current user"""
    data = request.get_json()
    printer_name = data.get('printer_name')
    printer_id = data.get('printer_id')
    
    current_user.printer_name = printer_name
    if printer_id is not None:
        current_user.printer_id = printer_id
    db.session.commit()
    
    return jsonify({
        'success': True,
        'printer_name': printer_name,
        'printer_id': printer_id
    })


# ============================================
# SERVER-SIDE PRINT QUEUE API
# ============================================

@app.route('/api/pending-prints')
@login_required
def get_pending_prints():
    """Get all pending prints from server-side queue"""
    pending = branch_filter(PendingPrint.query, PendingPrint).filter_by(status='pending').order_by(PendingPrint.created_at).all()
    return jsonify({
        'success': True,
        'pending_prints': [p.to_dict() for p in pending],
        'count': len(pending)
    })


@app.route('/api/pending-prints', methods=['POST'])
@login_required
def add_pending_print():
    """Add a new pending print to server-side queue"""
    data = request.get_json()
    
    pending = PendingPrint(
        order_id=data.get('order_id'),
        receipt_data=json.dumps(data.get('receipt_data', [])),
        copies=data.get('copies', 3),
        current_copy=data.get('current_copy', 1),
        status='pending',
        branch_id=get_default_branch_id()
    )
    
    db.session.add(pending)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'pending_print': pending.to_dict()
    })


@app.route('/api/pending-prints/<int:print_id>/complete', methods=['POST'])
@login_required
def complete_pending_print(print_id):
    """Mark a pending print as completed"""
    pending = PendingPrint.query.get_or_404(print_id)
    pending.status = 'completed'
    pending.printed_at = utc_now()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Print marked as completed'
    })


@app.route('/api/pending-prints/<int:print_id>/fail', methods=['POST'])
@login_required
def fail_pending_print(print_id):
    """Mark a pending print as failed and increment retry count"""
    data = request.get_json() or {}
    pending = PendingPrint.query.get_or_404(print_id)
    
    pending.retry_count += 1
    pending.error_message = data.get('error_message', 'Unknown error')
    
    # Mark as permanently failed after 5 retries
    if pending.retry_count >= 5:
        pending.status = 'failed'
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'retry_count': pending.retry_count,
        'status': pending.status
    })


@app.route('/api/pending-prints/<int:print_id>', methods=['DELETE'])
@login_required
def delete_pending_print(print_id):
    """Delete a pending print from queue"""
    pending = PendingPrint.query.get_or_404(print_id)
    db.session.delete(pending)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Print deleted from queue'
    })


@app.route('/api/pending-prints/clear', methods=['POST'])
@login_required
def clear_pending_prints():
    """Clear all pending prints"""
    bid = get_user_branch_id()
    query = PendingPrint.query.filter_by(status='pending')
    if bid is not None:
        query = query.filter(PendingPrint.branch_id == bid)
    query.delete()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'All pending prints cleared'
    })


# ============================================
# USB PRINTER API (Server-Side Printing)
# ============================================

@app.route('/api/usb-printer/status')
@login_required
def usb_printer_status():
    """Get USB printer status and availability"""
    if not USB_PRINTING_AVAILABLE:
        return jsonify({
            'success': True,
            'available': False,
            'connected': False,
            'message': 'USB printing not available (install python-escpos and pyusb)'
        })
    
    return jsonify({
        'success': True,
        'available': True,
        'connected': usb_printer.connected if usb_printer else False,
        'printer_info': usb_printer.printer_info if usb_printer else None
    })


@app.route('/api/usb-printer/devices')
@login_required
def list_usb_printers():
    """List available USB printers"""
    if not USB_PRINTING_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'USB printing not available',
            'devices': []
        })
    
    devices = USBPrinterManager.list_usb_devices()
    return jsonify({
        'success': True,
        'devices': devices
    })


@app.route('/api/usb-printer/connect', methods=['POST'])
@login_required
def connect_usb_printer():
    """Connect to USB printer"""
    if not USB_PRINTING_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'USB printing not available'
        })
    
    data = request.get_json() or {}
    vendor_id = data.get('vendor_id')
    product_id = data.get('product_id')
    
    success, message = usb_printer.connect(vendor_id, product_id)
    
    return jsonify({
        'success': success,
        'message': message,
        'connected': usb_printer.connected
    })


@app.route('/api/usb-printer/disconnect', methods=['POST'])
@login_required
def disconnect_usb_printer():
    """Disconnect USB printer"""
    if usb_printer:
        usb_printer.disconnect()
    
    return jsonify({
        'success': True,
        'message': 'Disconnected'
    })


@app.route('/api/usb-printer/test', methods=['POST'])
@login_required
def test_usb_print():
    """Test print on USB printer"""
    if not USB_PRINTING_AVAILABLE or not usb_printer:
        return jsonify({
            'success': False,
            'message': 'USB printer not available'
        })
    
    success, message = usb_printer.test_print()
    return jsonify({
        'success': success,
        'message': message
    })


@app.route('/api/usb-printer/print', methods=['POST'])
@login_required
def usb_print_receipt():
    """Print receipt on USB printer"""
    if not USB_PRINTING_AVAILABLE or not usb_printer:
        return jsonify({
            'success': False,
            'message': 'USB printer not available'
        })
    
    data = request.get_json()
    
    if 'order_data' in data:
        success, message = usb_printer.print_receipt(data['order_data'])
    elif 'raw_commands' in data:
        success, message = usb_printer.print_raw(data['raw_commands'])
    else:
        return jsonify({
            'success': False,
            'message': 'No print data provided'
        })
    
    return jsonify({
        'success': success,
        'message': message
    })


@app.route('/api/usb-printer/print-pending', methods=['POST'])
@login_required
def usb_print_pending():
    """Print all pending receipts using USB printer"""
    if not USB_PRINTING_AVAILABLE or not usb_printer or not usb_printer.connected:
        return jsonify({
            'success': False,
            'message': 'USB printer not connected'
        })
    
    pending = PendingPrint.query.filter_by(status='pending').order_by(PendingPrint.created_at).all()
    
    if not pending:
        return jsonify({
            'success': True,
            'message': 'No pending prints',
            'printed': 0
        })
    
    printed_count = 0
    failed_count = 0
    
    for item in pending:
        try:
            commands = json.loads(item.receipt_data)
            success, _ = usb_printer.print_raw(commands)
            
            if success:
                item.status = 'completed'
                item.printed_at = utc_now()
                printed_count += 1
            else:
                item.retry_count += 1
                if item.retry_count >= 5:
                    item.status = 'failed'
                failed_count += 1
                
        except Exception as e:
            item.retry_count += 1
            item.error_message = str(e)
            if item.retry_count >= 5:
                item.status = 'failed'
            failed_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'printed': printed_count,
        'failed': failed_count,
        'message': f'{printed_count} receipts printed, {failed_count} failed'
    })


# Reports
@app.route('/reports')
@login_required
@role_required('admin', 'manager')
def reports():
    return render_template('reports.html')

@app.route('/analytics')
@login_required
@role_required('admin', 'manager')
def analytics():
    """Comprehensive analytics dashboard with real data and percentages.
    Supports filtering by city_id, brand_id, and branch_id query parameters."""
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    two_weeks_ago = today - timedelta(days=14)
    month_start = today.replace(day=1)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    last_month_end = month_start - timedelta(days=1)
    
    # â”€â”€ Filter parameters (for owner/admin: can filter by city/brand/branch) â”€â”€
    filter_city_id = request.args.get('city_id', '', type=str).strip()
    filter_brand_id = request.args.get('brand_id', '', type=str).strip()
    filter_branch_id = request.args.get('branch_id', '', type=str).strip()
    
    # Query Pusat branch once for legacy NULL branch_id handling
    pusat_branch = Branch.query.filter_by(code='PUSAT').first()
    pusat_branch_id = pusat_branch.id if pusat_branch else None
    
    def _includes_pusat(branch_ids):
        """Check if branch_ids list includes Pusat (for legacy NULL data handling)."""
        return pusat_branch_id is not None and pusat_branch_id in branch_ids
    
    def analytics_filter(query):
        """Apply branch/city/brand filter for analytics. Regular users see their branch only."""
        bid = get_user_branch_id()
        if bid is not None:
            # Non-owner: locked to their branch
            return query.filter(Order.branch_id == bid)
        # Owner: apply optional filters
        if filter_branch_id and filter_branch_id.isdigit():
            target_id = int(filter_branch_id)
            # Also include legacy orders with NULL branch_id if filtering for Pusat
            if target_id == pusat_branch_id:
                return query.filter(db.or_(Order.branch_id == target_id, Order.branch_id.is_(None)))
            return query.filter(Order.branch_id == target_id)
        if filter_brand_id and filter_brand_id.isdigit():
            branch_ids = [id for (id,) in Branch.query.with_entities(Branch.id).filter_by(brand_id=int(filter_brand_id)).all()]
            if not branch_ids:
                return query.filter(Order.id == None)
            if _includes_pusat(branch_ids):
                return query.filter(db.or_(Order.branch_id.in_(branch_ids), Order.branch_id.is_(None)))
            return query.filter(Order.branch_id.in_(branch_ids))
        if filter_city_id and filter_city_id.isdigit():
            branch_ids = [id for (id,) in Branch.query.with_entities(Branch.id).filter_by(city_id=int(filter_city_id)).all()]
            if not branch_ids:
                return query.filter(Order.id == None)
            if _includes_pusat(branch_ids):
                return query.filter(db.or_(Order.branch_id.in_(branch_ids), Order.branch_id.is_(None)))
            return query.filter(Order.branch_id.in_(branch_ids))
        return query  # Owner with no filter = all data
    
    def analytics_expense_filter(query):
        """Apply branch/city/brand filter for expenses."""
        bid = get_user_branch_id()
        if bid is not None:
            return query.filter(Expense.branch_id == bid)
        if filter_branch_id and filter_branch_id.isdigit():
            target_id = int(filter_branch_id)
            if target_id == pusat_branch_id:
                return query.filter(db.or_(Expense.branch_id == target_id, Expense.branch_id.is_(None)))
            return query.filter(Expense.branch_id == target_id)
        if filter_brand_id and filter_brand_id.isdigit():
            branch_ids = [id for (id,) in Branch.query.with_entities(Branch.id).filter_by(brand_id=int(filter_brand_id)).all()]
            if not branch_ids:
                return query.filter(Expense.id == None)
            if _includes_pusat(branch_ids):
                return query.filter(db.or_(Expense.branch_id.in_(branch_ids), Expense.branch_id.is_(None)))
            return query.filter(Expense.branch_id.in_(branch_ids))
        if filter_city_id and filter_city_id.isdigit():
            branch_ids = [id for (id,) in Branch.query.with_entities(Branch.id).filter_by(city_id=int(filter_city_id)).all()]
            if not branch_ids:
                return query.filter(Expense.id == None)
            if _includes_pusat(branch_ids):
                return query.filter(db.or_(Expense.branch_id.in_(branch_ids), Expense.branch_id.is_(None)))
            return query.filter(Expense.branch_id.in_(branch_ids))
        return query
    
    # Helper: calculate growth percentage
    def growth_pct(current, previous):
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round(((current - previous) / previous) * 100, 1)
    
    # â”€â”€ Today vs Yesterday â”€â”€
    today_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) == today
    ).all()
    yesterday_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) == yesterday
    ).all()
    
    today_paid = [o for o in today_orders if o.payment and o.payment.status == 'paid']
    yesterday_paid = [o for o in yesterday_orders if o.payment and o.payment.status == 'paid']
    
    today_revenue = sum(o.total for o in today_paid)
    yesterday_revenue = sum(o.total for o in yesterday_paid)
    revenue_growth = growth_pct(today_revenue, yesterday_revenue)
    
    today_order_count = len(today_paid)
    yesterday_order_count = len(yesterday_paid)
    order_growth = growth_pct(today_order_count, yesterday_order_count)
    
    today_avg = today_revenue // today_order_count if today_order_count > 0 else 0
    yesterday_avg = yesterday_revenue // yesterday_order_count if yesterday_order_count > 0 else 0
    avg_growth = growth_pct(today_avg, yesterday_avg)
    
    today_cancelled = len([o for o in today_orders if o.status == 'cancelled'])
    today_cancel_rate = round(today_cancelled / len(today_orders) * 100, 1) if today_orders else 0
    
    # â”€â”€ This Week vs Last Week â”€â”€
    this_week_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) >= week_ago,
        db.func.date(Order.created_at) <= today
    ).all()
    last_week_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) >= two_weeks_ago,
        db.func.date(Order.created_at) < week_ago
    ).all()
    
    this_week_paid = [o for o in this_week_orders if o.payment and o.payment.status == 'paid']
    last_week_paid = [o for o in last_week_orders if o.payment and o.payment.status == 'paid']
    
    week_revenue = sum(o.total for o in this_week_paid)
    last_week_revenue = sum(o.total for o in last_week_paid)
    week_growth = growth_pct(week_revenue, last_week_revenue)
    
    # â”€â”€ This Month vs Last Month â”€â”€
    this_month_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) >= month_start,
        db.func.date(Order.created_at) <= today
    ).all()
    last_month_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) >= last_month_start,
        db.func.date(Order.created_at) <= last_month_end
    ).all()
    
    this_month_paid = [o for o in this_month_orders if o.payment and o.payment.status == 'paid']
    last_month_paid = [o for o in last_month_orders if o.payment and o.payment.status == 'paid']
    
    month_revenue = sum(o.total for o in this_month_paid)
    last_month_revenue = sum(o.total for o in last_month_paid)
    month_growth = growth_pct(month_revenue, last_month_revenue)
    
    # â”€â”€ Payment Method Breakdown (this month) â”€â”€
    payment_methods = {}
    for o in this_month_paid:
        method = o.payment.payment_method if o.payment else 'unknown'
        if method not in payment_methods:
            payment_methods[method] = {'count': 0, 'total': 0}
        payment_methods[method]['count'] += 1
        payment_methods[method]['total'] += o.total
    
    total_payment_count = sum(v['count'] for v in payment_methods.values())
    for method in payment_methods:
        payment_methods[method]['percentage'] = round(
            (payment_methods[method]['count'] / total_payment_count * 100), 1
        ) if total_payment_count > 0 else 0
    
    # â”€â”€ Order Source Breakdown (this month) â”€â”€
    source_breakdown = {}
    for o in this_month_paid:
        src = o.source or 'pos'
        if src not in source_breakdown:
            source_breakdown[src] = {'count': 0, 'total': 0}
        source_breakdown[src]['count'] += 1
        source_breakdown[src]['total'] += o.total
    
    total_source_count = sum(v['count'] for v in source_breakdown.values())
    for src in source_breakdown:
        source_breakdown[src]['percentage'] = round(
            (source_breakdown[src]['count'] / total_source_count * 100), 1
        ) if total_source_count > 0 else 0
    
    # â”€â”€ Order Type Breakdown (dine_in / takeaway / online) â”€â”€
    type_breakdown = {}
    for o in this_month_paid:
        otype = o.order_type or 'dine_in'
        if otype not in type_breakdown:
            type_breakdown[otype] = {'count': 0, 'total': 0}
        type_breakdown[otype]['count'] += 1
        type_breakdown[otype]['total'] += o.total
    
    total_type_count = sum(v['count'] for v in type_breakdown.values())
    for otype in type_breakdown:
        type_breakdown[otype]['percentage'] = round(
            type_breakdown[otype]['count'] / total_type_count * 100, 1
        ) if total_type_count > 0 else 0
    
    # â”€â”€ Category Performance (this month) â”€â”€
    category_perf = {}
    for o in this_month_paid:
        for item in o.items:
            cat_name = item.menu_item.category.name if item.menu_item and item.menu_item.category else 'Tanpa Kategori'
            if cat_name not in category_perf:
                category_perf[cat_name] = {'count': 0, 'qty': 0, 'total': 0}
            category_perf[cat_name]['count'] += 1
            category_perf[cat_name]['qty'] += item.quantity
            category_perf[cat_name]['total'] += item.subtotal
    
    total_cat_revenue = sum(v['total'] for v in category_perf.values())
    for cat in category_perf:
        category_perf[cat]['percentage'] = round(
            (category_perf[cat]['total'] / total_cat_revenue * 100), 1
        ) if total_cat_revenue > 0 else 0
    
    # Sort by revenue desc
    category_perf = dict(sorted(category_perf.items(), key=lambda x: x[1]['total'], reverse=True))
    
    # â”€â”€ Top 10 Menu Items (this month) â”€â”€
    top_items = {}
    for o in this_month_paid:
        for item in o.items:
            name = item.name
            if name not in top_items:
                top_items[name] = {'qty': 0, 'total': 0}
            top_items[name]['qty'] += item.quantity
            top_items[name]['total'] += item.subtotal
    
    top_items = dict(sorted(top_items.items(), key=lambda x: x[1]['qty'], reverse=True)[:10])
    total_item_qty = sum(v['qty'] for v in top_items.values())
    for name in top_items:
        top_items[name]['percentage'] = round(
            (top_items[name]['qty'] / total_item_qty * 100), 1
        ) if total_item_qty > 0 else 0
    
    # â”€â”€ Hourly Order Distribution (this week) â”€â”€
    hourly_dist = {h: 0 for h in range(24)}
    for o in this_week_paid:
        if o.created_at:
            hourly_dist[o.created_at.hour] += 1
    
    peak_hour = max(hourly_dist, key=hourly_dist.get) if any(hourly_dist.values()) else 12
    
    # â”€â”€ Daily Revenue Trend (last 30 days) â”€â”€
    thirty_days_ago = today - timedelta(days=30)
    month_all_orders = analytics_filter(Order.query).filter(
        db.func.date(Order.created_at) >= thirty_days_ago,
        db.func.date(Order.created_at) <= today
    ).all()
    
    daily_trend = {}
    for o in month_all_orders:
        if o.payment and o.payment.status == 'paid':
            day_str = o.created_at.strftime('%Y-%m-%d')
            if day_str not in daily_trend:
                daily_trend[day_str] = {'revenue': 0, 'orders': 0}
            daily_trend[day_str]['revenue'] += o.total
            daily_trend[day_str]['orders'] += 1
    
    # Fill missing days with 0
    daily_labels = []
    daily_revenues = []
    daily_orders_list = []
    current_day = thirty_days_ago
    while current_day <= today:
        day_str = current_day.strftime('%Y-%m-%d')
        daily_labels.append(current_day.strftime('%d/%m'))
        daily_revenues.append(daily_trend.get(day_str, {}).get('revenue', 0))
        daily_orders_list.append(daily_trend.get(day_str, {}).get('orders', 0))
        current_day += timedelta(days=1)
    
    # â”€â”€ Expenses this month â”€â”€
    expense_query = analytics_expense_filter(Expense.query.filter(
        db.func.date(Expense.date) >= month_start,
        db.func.date(Expense.date) <= today
    ))
    total_expenses = sum(e.amount for e in expense_query.all())
    net_profit = month_revenue - total_expenses
    profit_margin = round((net_profit / month_revenue * 100), 1) if month_revenue > 0 else 0
    
    return render_template('analytics.html',
        today_revenue=today_revenue,
        yesterday_revenue=yesterday_revenue,
        revenue_growth=revenue_growth,
        today_order_count=today_order_count,
        order_growth=order_growth,
        today_avg=today_avg,
        avg_growth=avg_growth,
        today_cancel_rate=today_cancel_rate,
        week_revenue=week_revenue,
        week_growth=week_growth,
        month_revenue=month_revenue,
        month_growth=month_growth,
        month_orders=len(this_month_paid),
        payment_methods=payment_methods,
        source_breakdown=source_breakdown,
        type_breakdown=type_breakdown,
        category_perf=category_perf,
        top_items=top_items,
        hourly_dist=hourly_dist,
        peak_hour=peak_hour,
        daily_labels=daily_labels,
        daily_revenues=daily_revenues,
        daily_orders_list=daily_orders_list,
        total_expenses=total_expenses,
        net_profit=net_profit,
        profit_margin=profit_margin,
        filter_city_id=filter_city_id,
        filter_brand_id=filter_brand_id,
        filter_branch_id=filter_branch_id,
        cities=City.query.order_by(City.name).all(),
        brands=Brand.query.order_by(Brand.name).all(),
        branches=Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    )

@app.route('/reports/income')
@login_required
@role_required('admin', 'manager')
def income_report():
    # Get date range from query params
    start_date_str = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    # Parse dates properly
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
    
    orders = branch_filter(Order.query, Order).filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    total_income = sum(o.total for o in paid_orders)
    total_orders = len(paid_orders)
    
    # Group by date
    daily_income = {}
    for order in paid_orders:
        date_str = order.created_at.strftime('%Y-%m-%d')
        if date_str not in daily_income:
            daily_income[date_str] = {'income': 0, 'orders': 0}
        daily_income[date_str]['income'] += order.total
        daily_income[date_str]['orders'] += 1
    
    return render_template('reports/income.html',
                         orders=paid_orders,
                         total_income=total_income,
                         total_orders=total_orders,
                         daily_income=daily_income,
                         start_date=start_date_str,
                         end_date=end_date_str)

@app.route('/reports/export/pdf')
@login_required
@role_required('admin', 'manager')
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as PDFTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    orders = branch_filter(Order.query, Order).filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date + ' 23:59:59'
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph(f"Laporan Penjualan", styles['Heading1']))
    elements.append(Paragraph(f"Periode: {start_date} - {end_date}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Tanggal', 'Order ID', 'Total']]
    for i, order in enumerate(paid_orders, 1):
        data.append([
            str(i),
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.order_number,
            f"Rp {order.total:,}".replace(',', '.')
        ])
    
    # Total row
    total = sum(o.total for o in paid_orders)
    data.append(['', '', 'TOTAL', f"Rp {total:,}".replace(',', '.')])
    
    table = PDFTable(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'laporan_{start_date}_{end_date}.pdf',
        mimetype='application/pdf'
    )

@app.route('/reports/export/excel')
@login_required
@role_required('admin', 'manager')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    orders = branch_filter(Order.query, Order).filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date + ' 23:59:59'
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    
    # Header
    ws['A1'] = 'Laporan Penjualan'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Periode: {start_date} - {end_date}'
    
    # Column headers
    headers = ['No', 'Tanggal', 'Order ID', 'Customer', 'Subtotal', 'Total', 'Payment Method']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, order in enumerate(paid_orders, 5):
        ws.cell(row=row, column=1, value=row-4)
        ws.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=order.order_number)
        ws.cell(row=row, column=4, value=order.customer_name or '-')
        ws.cell(row=row, column=5, value=order.subtotal)
        ws.cell(row=row, column=6, value=order.total)
        ws.cell(row=row, column=7, value=order.payment.payment_method if order.payment else '-')
    
    # Total
    total_row = len(paid_orders) + 5
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(o.total for o in paid_orders)).font = Font(bold=True)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'laporan_{start_date}_{end_date}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Payment page with custom design
@app.route('/payment/<int:order_id>')
@login_required
def payment_page(order_id):
    order = db.session.get(Order, order_id)
    if not order:
        flash('Pesanan tidak ditemukan', 'error')
        return redirect(url_for('orders'))
    
    # Get snap token from payment or generate new one
    snap_token = None
    if order.payment and order.payment.snap_token:
        snap_token = order.payment.snap_token
    elif order.payment and order.payment.status == 'pending':
        # Generate snap token if not exists
        try:
            import midtransclient
            
            snap = midtransclient.Snap(
                is_production=app.config.get('MIDTRANS_IS_PRODUCTION', False),
                server_key=app.config.get('MIDTRANS_SERVER_KEY', ''),
                client_key=app.config.get('MIDTRANS_CLIENT_KEY', '')
            )
            
            param = {
                "transaction_details": {
                    "order_id": f"DTO-{order.id}-{int(datetime.now().timestamp())}",
                    "gross_amount": int(order.total)
                },
                "customer_details": {
                    "first_name": current_user.full_name or current_user.username,
                    "email": current_user.email or f"{current_user.username}@dapoerterasobor.com"
                },
                "item_details": [{
                    "id": str(item.menu_item_id),
                    "price": int(item.price),
                    "quantity": item.quantity,
                    "name": item.menu_item.name[:50]
                } for item in order.items]
            }
            
            transaction = snap.create_transaction(param)
            snap_token = transaction.get('token')
            
            # Save snap token
            order.payment.snap_token = snap_token
            db.session.commit()
        except Exception as e:
            print(f"Error generating snap token: {e}")
    
    return render_template('payment.html', 
                         order=order, 
                         snap_token=snap_token or '',
                         auto_pay=bool(snap_token),
                         config=app.config)

# API to update payment status from frontend
@app.route('/api/payment/<int:order_id>/status', methods=['POST'])
@login_required
def update_payment_status(order_id):
    order = db.session.get(Order, order_id)
    if not order or not order.payment:
        return jsonify({'error': 'Order not found'}), 404
    
    data = request.json
    status = data.get('status', 'pending')
    
    if status == 'paid':
        order.payment.status = 'paid'
        order.payment.paid_at = utc_now()
        order.status = 'completed'
    elif status == 'pending':
        order.payment.status = 'pending'
    elif status == 'failed':
        order.payment.status = 'failed'
    
    db.session.commit()
    return jsonify({'success': True, 'status': order.payment.status})

# Order management
@app.route('/orders')
@login_required
@role_required('admin', 'manager', 'kasir')
def orders():
    status_filter = request.args.get('status', 'all')
    
    query = branch_filter(Order.query, Order).order_by(Order.created_at.desc())
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    orders = query.limit(100).all()
    
    return render_template('orders.html', orders=orders, status_filter=status_filter)

# Error handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('errors/500.html'), 500

# Admin Reset Database
@app.route('/admin/reset-database', methods=['POST'])
@login_required
@role_required('admin')
def reset_database():
    """Reset all transactional data (orders, payments, carts) but keep menu, users, tables"""
    try:
        # Clear cart items first (foreign key)
        CartItem.query.delete()
        Cart.query.delete()
        
        # Clear order items (foreign key)
        OrderItem.query.delete()
        
        # Clear payments (foreign key)
        Payment.query.delete()
        
        # Clear orders
        Order.query.delete()
        
        # Clear income records
        Income.query.delete()
        
        # Reset table status
        Table.query.update({Table.status: 'available'})
        
        db.session.commit()
        flash('Database berhasil direset! Semua data pesanan, pembayaran, dan keranjang telah dihapus.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saat reset database: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

# ========== NOTIFICATION SYSTEM ==========
@app.route('/api/notifications')
@login_required
def api_get_notifications():
    """Get notifications for current user"""
    notifications = branch_filter(Notification.query, Notification).filter(
        (Notification.user_id == current_user.id) | (Notification.user_id == None)
    ).order_by(Notification.created_at.desc()).limit(50).all()
    
    unread_count = branch_filter(Notification.query, Notification).filter(
        ((Notification.user_id == current_user.id) | (Notification.user_id == None)),
        Notification.is_read == False
    ).count()
    
    return jsonify({
        'notifications': [n.to_dict() for n in notifications],
        'unread_count': unread_count
    })


@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def api_mark_notification_read(notification_id):
    """Mark a notification as read"""
    notification = Notification.query.get_or_404(notification_id)
    notification.is_read = True
    notification.read_at = utc_now()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_mark_all_notifications_read():
    """Mark all notifications as read for current user"""
    Notification.query.filter(
        ((Notification.user_id == current_user.id) | (Notification.user_id == None)),
        Notification.is_read == False
    ).update({'is_read': True, 'read_at': utc_now()}, synchronize_session=False)
    db.session.commit()
    return jsonify({'success': True})


def create_notification(type, title, message, user_id=None, data=None):
    """Helper function to create a notification"""
    import json
    notification = Notification(
        type=type,
        title=title,
        message=message,
        user_id=user_id,
        data=json.dumps(data) if data else None,
        branch_id=get_default_branch_id()
    )
    db.session.add(notification)
    db.session.commit()
    return notification



# ============================================
# FOOD DELIVERY PLATFORM INTEGRATION (Production)
# ============================================
# Supports real integration with:
# - GrabFood (partner-api.grab.com) - OAuth2 client_credentials, HMAC-SHA256
# - GoFood/GoBiz (api.gobiz.co.id) - OAuth2 client_credentials, event-based webhooks
# - ShopeeFood (partner.shopeemobile.com) - partner_key HMAC-SHA256 signing
#
# Two integration modes per platform:
# 1. WEBHOOK MODE: Platform sends orders to our endpoint (real-time)
# 2. POLLING MODE: We periodically fetch new orders from platform API (fallback)
# ============================================

# --- Platform API Configuration ---
PLATFORM_CONFIG = {
    'grabfood': {
        'name': 'GrabFood',
        'oauth_token_url': 'https://partner-api.grab.com/grabid/v1/oauth2/token',
        'oauth_token_url_sandbox': 'https://partner-api.stg-myteksi.com/grabid/v1/oauth2/token',
        'api_base': 'https://partner-api.grab.com',
        'api_base_sandbox': 'https://partner-api.stg-myteksi.com',
        'scope': 'food.partner_api',
        'sig_header': 'X-Grab-Signature',
        'sig_algo': 'sha256',
        'webhook_events': ['order.placed', 'order.cancelled', 'order.completed'],
    },
    'gofood': {
        'name': 'GoFood',
        'oauth_token_url': 'https://accounts.go-jek.com/oauth2/token',
        'oauth_token_url_sandbox': 'https://integration-goauth.gojekapi.com/oauth2/token',
        'api_base': 'https://api.gobiz.co.id',
        'api_base_sandbox': 'https://api.partner-sandbox.gobiz.co.id',
        'scope': 'gofood:order:read gofood:order:write gofood:catalog:read',
        'sig_header': 'X-Callback-Token',
        'sig_algo': 'token',
        'webhook_events': ['gofood.order.awaiting_merchant_acceptance', 'gofood.order.merchant_accepted',
                           'gofood.order.cancelled', 'gofood.order.completed',
                           'gofood.order.driver_otw_pickup', 'gofood.order.driver_arrived'],
    },
    'shopeefood': {
        'name': 'ShopeeFood',
        'api_base': 'https://partner.shopeemobile.com',
        'api_base_sandbox': 'https://partner.test-stable.shopeemobile.com',
        'sig_header': 'Authorization',
        'sig_algo': 'sha256',
        'webhook_events': ['ORDER_STATUS_UPDATE', 'ORDER_CREATE', 'ORDER_CANCEL'],
    }
}

# In-memory token cache (per platform)
_oauth_tokens = {}


def get_platform_config(platform):
    """Get configuration for a platform from environment/app config"""
    prefix = platform.upper()
    is_sandbox = app.config.get(f'{prefix}_SANDBOX', True)
    pcfg = PLATFORM_CONFIG.get(platform, {})
    
    return {
        'client_id': os.environ.get(f'{prefix}_CLIENT_ID', ''),
        'client_secret': os.environ.get(f'{prefix}_CLIENT_SECRET', ''),
        'webhook_secret': os.environ.get(f'{prefix}_WEBHOOK_SECRET', ''),
        'partner_id': os.environ.get(f'{prefix}_PARTNER_ID', ''),
        'partner_key': os.environ.get(f'{prefix}_PARTNER_KEY', ''),
        'merchant_id': os.environ.get(f'{prefix}_MERCHANT_ID', ''),
        'store_id': os.environ.get(f'{prefix}_STORE_ID', ''),
        'store_branch_map': os.environ.get(f'{prefix}_STORE_BRANCH_MAP', ''),
        'is_sandbox': is_sandbox,
        'api_base': pcfg.get('api_base_sandbox') if is_sandbox else pcfg.get('api_base', ''),
        'token_url': pcfg.get('oauth_token_url_sandbox') if is_sandbox else pcfg.get('oauth_token_url', ''),
        'scope': pcfg.get('scope', ''),
    }


# ---- OAuth2 Token Management ----

def get_oauth_token(platform):
    """
    Get a valid OAuth2 access token for a platform.
    Uses client_credentials grant (GrabFood, GoFood).
    Caches tokens until expiry.
    """
    import requests as http_requests
    
    # Check cache
    cached = _oauth_tokens.get(platform)
    if cached and cached['expires_at'] > datetime.now():
        return cached['access_token']
    
    cfg = get_platform_config(platform)
    token_url = cfg.get('token_url', '')
    client_id = cfg.get('client_id', '')
    client_secret = cfg.get('client_secret', '')
    
    if not token_url or not client_id or not client_secret:
        return None
    
    try:
        if platform == 'grabfood':
            # GrabFood: POST with form data
            resp = http_requests.post(token_url, data={
                'client_id': client_id,
                'client_secret': client_secret,
                'grant_type': 'client_credentials',
                'scope': cfg.get('scope', 'food.partner_api'),
            }, headers={'Content-Type': 'application/x-www-form-urlencoded'}, timeout=15)
            
        elif platform == 'gofood':
            # GoFood/GoBiz: POST with Basic auth
            resp = http_requests.post(token_url,
                auth=(client_id, client_secret),
                data={
                    'grant_type': 'client_credentials',
                    'scope': cfg.get('scope', 'gofood:order:read gofood:order:write'),
                },
                headers={'Content-Type': 'application/x-www-form-urlencoded'},
                timeout=15)
        else:
            return None
        
        if resp.status_code == 200:
            data = resp.json()
            expires_in = data.get('expires_in', 3600)
            _oauth_tokens[platform] = {
                'access_token': data['access_token'],
                'expires_at': datetime.now() + timedelta(seconds=expires_in - 60),
            }
            return data['access_token']
        else:
            app.logger.error(f'OAuth token error for {platform}: {resp.status_code} {resp.text}')
            return None
    except Exception as e:
        app.logger.error(f'OAuth token request failed for {platform}: {e}')
        return None


# ---- Platform-Specific Signature Verification ----

def verify_webhook_signature(platform, request_obj):
    """
    Verify webhook signature using the platform's actual method.
    
    GrabFood: HMAC-SHA256(body, secret) â†’ X-Grab-Signature header
    GoFood:   Direct token comparison â†’ X-Callback-Token header
    ShopeeFood: HMAC-SHA256(base_string, partner_key) â†’ Authorization header
    
    Returns: (is_valid, sig_status, detail)
    """
    cfg = get_platform_config(platform)
    pcfg = PLATFORM_CONFIG.get(platform, {})
    
    if platform == 'grabfood':
        secret = cfg.get('webhook_secret', '')
        if not secret:
            return True, 'skipped', 'No GRABFOOD_WEBHOOK_SECRET configured'
        
        sig_header = request_obj.headers.get('X-Grab-Signature', '')
        if not sig_header:
            # Fallback for direct secret
            fallback = request_obj.headers.get('X-Webhook-Secret', '')
            if fallback and hmac.compare_digest(fallback, secret):
                return True, 'valid', 'Direct secret match (fallback)'
            return False, 'missing', 'Missing X-Grab-Signature header'
        
        body = request_obj.get_data()
        computed = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
        if hmac.compare_digest(sig_header, computed):
            return True, 'valid', 'HMAC-SHA256 signature valid'
        return False, 'invalid', 'X-Grab-Signature HMAC-SHA256 mismatch'
    
    elif platform == 'gofood':
        # GoFood uses X-Callback-Token â€” a static token comparison, NOT HMAC of body
        secret = cfg.get('webhook_secret', '')
        if not secret:
            return True, 'skipped', 'No GOFOOD_WEBHOOK_SECRET configured'
        
        callback_token = request_obj.headers.get('X-Callback-Token', '')
        if not callback_token:
            fallback = request_obj.headers.get('X-Webhook-Secret', '')
            if fallback and hmac.compare_digest(fallback, secret):
                return True, 'valid', 'Direct secret match (fallback)'
            return False, 'missing', 'Missing X-Callback-Token header'
        
        if hmac.compare_digest(callback_token, secret):
            return True, 'valid', 'X-Callback-Token verified'
        return False, 'invalid', 'X-Callback-Token mismatch'
    
    elif platform == 'shopeefood':
        # ShopeeFood: HMAC-SHA256(partner_key, base_string)
        # base_string = partner_id + api_path + timestamp + access_token + shop_id
        partner_key = cfg.get('partner_key', '') or cfg.get('webhook_secret', '')
        if not partner_key:
            return True, 'skipped', 'No SHOPEEFOOD_PARTNER_KEY configured'
        
        sig_header = request_obj.headers.get('Authorization', '')
        if not sig_header:
            sig_header = request_obj.headers.get('X-Shopee-Hmac-SHA256', '')
        if not sig_header:
            fallback = request_obj.headers.get('X-Webhook-Secret', '')
            if fallback and hmac.compare_digest(fallback, partner_key):
                return True, 'valid', 'Direct secret match (fallback)'
            return False, 'missing', 'Missing Authorization/X-Shopee-Hmac-SHA256 header'
        
        # Try HMAC verification of body
        body = request_obj.get_data()
        computed = hmac.new(partner_key.encode(), body, hashlib.sha256).hexdigest()
        if hmac.compare_digest(sig_header, computed):
            return True, 'valid', 'HMAC-SHA256 signature valid'
        
        # Also try direct comparison (some setups use static token)
        if hmac.compare_digest(sig_header, partner_key):
            return True, 'valid', 'Direct key match'
        
        return False, 'invalid', 'Shopee signature mismatch'
    
    return False, 'invalid', f'Unknown platform: {platform}'


def log_webhook_attempt(platform, request_obj, result, error_detail=None, ext_order_id=None):
    """Log webhook attempt for auditing and debugging signature failures"""
    try:
        body = request_obj.get_data()
        body_hash = hashlib.sha256(body).hexdigest() if body else None
        
        relevant_headers = {}
        for h in ['Content-Type', 'X-Grab-Signature', 'X-Callback-Token',
                   'X-Shopee-Hmac-SHA256', 'X-Webhook-Secret', 'X-Forwarded-For',
                   'User-Agent', 'X-Branch-Code', 'X-Store-ID', 'Authorization']:
            val = request_obj.headers.get(h)
            if val:
                if any(s in h.lower() for s in ('signature', 'secret', 'token', 'hmac', 'authorization')):
                    relevant_headers[h] = val[:4] + '****' if len(val) > 4 else '****'
                else:
                    relevant_headers[h] = val
        
        log_entry = WebhookLog(
            platform=platform,
            request_ip=request_obj.headers.get('X-Forwarded-For', request_obj.remote_addr),
            request_headers=json.dumps(relevant_headers),
            request_body_hash=body_hash,
            signature_provided=(request_obj.headers.get('X-Grab-Signature', '') or
                               request_obj.headers.get('X-Callback-Token', '') or
                               request_obj.headers.get('X-Shopee-Hmac-SHA256', '') or
                               request_obj.headers.get('Authorization', ''))[:8] + '****',
            result=result,
            error_detail=error_detail,
            external_order_id=ext_order_id
        )
        db.session.add(log_entry)
        db.session.commit()
    except Exception:
        db.session.rollback()


# ---- Platform-Specific Payload Extraction ----

def extract_grabfood_order(data):
    """
    Extract order from real GrabFood webhook payload.
    
    GrabFood SubmitOrder payload structure (from official SDK):
    {
      "orderID": "...",
      "shortOrderNumber": "123",
      "merchantID": "...",
      "partnerMerchantID": "...",
      "paymentType": "CASHLESS",
      "orderTime": "2025-...",
      "currency": {"code": "IDR", "symbol": "Rp", "exponent": 0},
      "items": [{"id": "...", "grabItemID": "...", "quantity": 1, "price": 35000,
                 "tax": 0, "specifications": "extra spicy",
                 "modifiers": [{"id": "...", "name": "...", "quantity": 1, "price": 5000}]}],
      "price": {"subtotal": 35000, "tax": 0, "merchantChargeFee": 0,
                "grabFundPromo": 0, "merchantFundPromo": 0, "eaterPayment": 35000},
      "receiver": {"name": "John", "phones": ["+62812..."], "address": {...}},
      "featureFlags": {"orderAcceptedType": "AUTO" | "MANUAL"}
    }
    """
    result = {'external_id': '', 'customer_name': '', 'items': [], 'notes': '',
              'store_id': '', 'total': 0}
    
    # Order ID: real field is "orderID", fallback to "order_id"
    result['external_id'] = (data.get('orderID') or data.get('order_id') or
                              data.get('shortOrderNumber', ''))
    
    # Store/Merchant ID for branch routing
    result['store_id'] = data.get('merchantID') or data.get('merchant_id') or data.get('partnerMerchantID', '')
    
    # Customer from receiver object (real GrabFood structure)
    receiver = data.get('receiver', {})
    if isinstance(receiver, dict):
        result['customer_name'] = receiver.get('name', '')
    # Fallback to eater (older format)
    if not result['customer_name']:
        eater = data.get('eater', {})
        if isinstance(eater, dict):
            result['customer_name'] = eater.get('name', '')
    if not result['customer_name']:
        result['customer_name'] = data.get('customer_name', '')
    
    # Items from real structure
    raw_items = data.get('items', []) or data.get('order_items', [])
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        
        # In real GrabFood, "id" is the partner's externalID, not the item name
        # The item name is in the linked menu, so we use 'id' to match
        item_id = item.get('id', '')
        grab_item_id = item.get('grabItemID') or item.get('grab_item_id', '')
        item_name = item.get('name', '') or item_id  # Some payloads include name
        
        # Price in real API is in minor unit (cents), but for IDR exponent=0 so it's normal
        currency = data.get('currency', {})
        exponent = currency.get('exponent', 0) if isinstance(currency, dict) else 0
        
        raw_price = int(item.get('price', 0) or 0)
        item_price = raw_price if exponent == 0 else int(raw_price / (10 ** exponent))
        
        item_qty = int(item.get('quantity', 1) or 1)
        item_notes = item.get('specifications', '') or item.get('notes', '')
        
        # Process modifiers/add-ons
        modifiers = item.get('modifiers', []) or []
        modifier_notes = []
        for mod in modifiers:
            if isinstance(mod, dict):
                mod_name = mod.get('name', '')
                if mod_name:
                    modifier_notes.append(mod_name)
        if modifier_notes:
            item_notes = (item_notes + ' | ' + ', '.join(modifier_notes)).strip(' | ')
        
        result['items'].append({
            'external_id': item_id,
            'grab_item_id': grab_item_id,
            'name': item_name,
            'quantity': item_qty,
            'price': item_price,
            'notes': item_notes,
        })
    
    # Total from price object
    price_obj = data.get('price', {})
    if isinstance(price_obj, dict):
        result['total'] = int(price_obj.get('subtotal', 0) or 0)
    
    # Notes
    result['notes'] = data.get('specialInstructions', '') or data.get('notes', '') or data.get('special_instructions', '')
    
    return result


def extract_gofood_order(data):
    """
    Extract order from real GoFood/GoBiz webhook payload.
    
    GoFood notification payload structure (from official docs):
    {
      "header": {
        "version": 1,
        "timestamp": "2025-...",
        "event_name": "gofood.order.awaiting_merchant_acceptance",
        "event_id": "uuid"
      },
      "body": {
        "service_type": "gofood",
        "outlet": {"id": "M008272", "external_outlet_id": "1", "name": "..."},
        "customer": {"id": "...", "name": "Jane Doe", "phone_number": "..."},
        "order": {
          "order_id": "OF12345",
          "items": [{"item_id": "I98765", "name": "Nasi Goreng", "quantity": 2,
                     "price": 25000, "special_instructions": "No chili"}],
          "total_price": 58000,
          "note": "Ring the bell"
        }
      }
    }
    """
    result = {'external_id': '', 'customer_name': '', 'items': [], 'notes': '',
              'store_id': '', 'total': 0, 'event_name': ''}
    
    # GoFood has header + body structure
    header = data.get('header', {})
    body = data.get('body', data)  # Fallback to top-level if no wrapper
    
    result['event_name'] = header.get('event_name', '') if isinstance(header, dict) else ''
    
    # Order data is nested in body.order
    order_data = body.get('order', body) if isinstance(body, dict) else data
    
    # Order ID
    if isinstance(order_data, dict):
        result['external_id'] = (order_data.get('order_id') or order_data.get('id') or
                                  data.get('order_id') or data.get('id', ''))
    else:
        result['external_id'] = data.get('order_id') or data.get('id', '')
    
    # Store/Outlet ID for branch routing
    outlet = body.get('outlet', {}) if isinstance(body, dict) else {}
    if isinstance(outlet, dict):
        result['store_id'] = outlet.get('id') or outlet.get('external_outlet_id', '')
    
    # Customer
    customer = body.get('customer', {}) if isinstance(body, dict) else {}
    if isinstance(customer, dict):
        result['customer_name'] = customer.get('name', '')
    if not result['customer_name']:
        result['customer_name'] = data.get('customer_name', '')
    
    # Items
    raw_items = []
    if isinstance(order_data, dict):
        raw_items = order_data.get('items', []) or []
    if not raw_items:
        raw_items = data.get('items', []) or data.get('order_items', [])
    
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        result['items'].append({
            'external_id': item.get('item_id') or item.get('id', ''),
            'name': item.get('name') or item.get('food_name', 'Item'),
            'quantity': int(item.get('quantity') or item.get('qty', 1)),
            'price': int(item.get('price') or item.get('unit_price', 0)),
            'notes': item.get('special_instructions') or item.get('notes') or item.get('variant_name', ''),
        })
    
    # Total
    if isinstance(order_data, dict):
        result['total'] = int(order_data.get('total_price', 0) or 0)
    if not result['total']:
        result['total'] = int(data.get('total_price', 0) or 0)
    
    # Notes
    if isinstance(order_data, dict):
        result['notes'] = order_data.get('note', '') or order_data.get('notes', '')
    if not result['notes']:
        result['notes'] = data.get('notes', '') or data.get('customer_note', '')
    
    return result


def extract_shopeefood_order(data):
    """
    Extract order from real ShopeeFood/Shopee Open Platform webhook payload.
    
    Shopee push notification structure:
    {
      "shop_id": 123456,
      "code": 0,  (0 = success)
      "data": {
        "ordersn": "220421ABCDEF",
        "order_status": "READY_TO_SHIP",
        "buyer_username": "buyer123",
        "item_list": [
          {"item_id": 123, "item_name": "Nasi Goreng", "model_quantity_purchased": 2,
           "model_original_price": 35000}
        ],
        "total_amount": 70000,
        "message_to_seller": "Extra pedas"
      }
    }
    
    Also supports simpler format for ShopeeFood-specific integration:
    {
      "order_code": "SPF12345",
      "buyer": {"name": "..."},
      "items": [{"name": "...", "quantity": 1, "price": 35000}]
    }
    """
    result = {'external_id': '', 'customer_name': '', 'items': [], 'notes': '',
              'store_id': '', 'total': 0}
    
    # Check for Shopee Open Platform nested "data" structure
    shopee_data = data.get('data', data)  # Use "data" sub-object if present
    
    # Store ID
    result['store_id'] = str(data.get('shop_id', '') or data.get('store_id', ''))
    
    # Order ID
    if isinstance(shopee_data, dict):
        result['external_id'] = (shopee_data.get('ordersn') or shopee_data.get('order_sn') or
                                  shopee_data.get('order_code') or data.get('order_code') or
                                  data.get('order_id', ''))
    else:
        result['external_id'] = data.get('order_code') or data.get('order_id', '')
    
    # Customer
    buyer = data.get('buyer', {})
    if isinstance(buyer, dict):
        result['customer_name'] = buyer.get('name') or buyer.get('username', '')
    if not result['customer_name'] and isinstance(shopee_data, dict):
        result['customer_name'] = shopee_data.get('buyer_username', '')
    if not result['customer_name']:
        result['customer_name'] = data.get('customer_name', '')
    
    # Items - Shopee Open Platform uses "item_list", ShopeeFood uses "items"
    raw_items = []
    if isinstance(shopee_data, dict):
        raw_items = shopee_data.get('item_list', []) or []
    if not raw_items:
        raw_items = data.get('items', []) or data.get('order_items', [])
    
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        result['items'].append({
            'external_id': str(item.get('item_id') or item.get('id', '')),
            'name': item.get('item_name') or item.get('name', 'Item'),
            'quantity': int(item.get('model_quantity_purchased') or item.get('quantity') or item.get('amount', 1)),
            'price': int(item.get('model_original_price') or item.get('price') or item.get('original_price', 0)),
            'notes': item.get('special_instructions', '') or item.get('notes', ''),
        })
    
    # Total
    if isinstance(shopee_data, dict):
        result['total'] = int(shopee_data.get('total_amount', 0) or 0)
    if not result['total']:
        result['total'] = int(data.get('total_amount', 0) or 0)
    
    # Notes
    if isinstance(shopee_data, dict):
        result['notes'] = shopee_data.get('message_to_seller', '') or shopee_data.get('note', '')
    if not result['notes']:
        result['notes'] = data.get('message_to_seller', '') or data.get('note', '') or data.get('notes', '')
    
    return result


# ---- Branch Resolution ----

def resolve_branch_from_platform(platform, parsed_data, request_obj):
    """
    Resolve branch using platform's store_id from payload (NOT X-Branch-Code).
    
    Priority:
    1. Match store_id from payload to configured PLATFORM_STORE_BRANCH_MAP
    2. Match X-Store-ID header
    3. Fallback X-Branch-Code header (backward compat)
    4. Default to Pusat
    """
    branch_id = None
    store_id = parsed_data.get('store_id', '')
    store_branch_map_str = app.config.get(f'{platform.upper()}_STORE_BRANCH_MAP', '')
    
    # 1) Match store_id from payload
    if store_id and store_branch_map_str:
        for mapping in store_branch_map_str.split(','):
            parts = mapping.strip().split(':')
            if len(parts) == 2 and parts[0].strip() == str(store_id):
                branch = Branch.query.filter_by(code=parts[1].strip(), is_active=True).first()
                if branch:
                    branch_id = branch.id
                    break
    
    # 2) X-Store-ID header
    if not branch_id:
        header_store_id = request_obj.headers.get('X-Store-ID', '')
        if header_store_id and store_branch_map_str:
            for mapping in store_branch_map_str.split(','):
                parts = mapping.strip().split(':')
                if len(parts) == 2 and parts[0].strip() == header_store_id:
                    branch = Branch.query.filter_by(code=parts[1].strip(), is_active=True).first()
                    if branch:
                        branch_id = branch.id
                        break
    
    # 3) X-Branch-Code fallback
    if not branch_id:
        branch_code = request_obj.headers.get('X-Branch-Code', '')
        if branch_code:
            branch = Branch.query.filter_by(code=branch_code, is_active=True).first()
            if branch:
                branch_id = branch.id
    
    # 4) Default Pusat
    if not branch_id:
        pusat = Branch.query.filter_by(code='PUSAT').first()
        if pusat:
            branch_id = pusat.id
    
    return branch_id


# ---- Core Order Processing ----

def process_platform_order(platform, data, branch_id=None):
    """
    Process an incoming order from a food delivery platform.
    
    Uses platform-specific extractors, creates order with proper status flow:
    received â†’ accepted â†’ printed
    
    Printing is DEFERRED to the print worker (not inline in webhook).
    """
    # Extract with platform-specific parser
    extractors = {
        'grabfood': extract_grabfood_order,
        'gofood': extract_gofood_order,
        'shopeefood': extract_shopeefood_order,
    }
    extractor = extractors.get(platform)
    if not extractor:
        return None, 'Platform tidak dikenal'
    
    parsed = extractor(data)
    external_id = parsed['external_id']
    customer_name = parsed['customer_name']
    items_data = parsed['items']
    notes = parsed['notes']
    
    if not external_id:
        return None, 'Order ID tidak ditemukan dalam payload'
    
    # Idempotency check
    existing = ExternalOrder.query.filter_by(
        platform=platform, external_order_id=str(external_id)
    ).first()
    if existing:
        return existing, 'Order sudah diproses sebelumnya'
    
    # Create ExternalOrder record
    ext_order = ExternalOrder(
        platform=platform,
        external_order_id=str(external_id),
        raw_data=json.dumps(data),
        status='received',
        branch_id=branch_id
    )
    db.session.add(ext_order)
    db.session.flush()
    
    try:
        # Generate order number
        platform_prefix = {'grabfood': 'GRB', 'gofood': 'GOF', 'shopeefood': 'SPF'}
        random_suffix = random.randint(100, 999)
        order_number = f"{platform_prefix.get(platform, 'EXT')}{datetime.now().strftime('%Y%m%d%H%M%S')}{random_suffix}"
        
        # Fallback customer name
        if not customer_name:
            platform_labels = {'grabfood': 'GrabFood', 'gofood': 'GoFood', 'shopeefood': 'ShopeeFood'}
            suffix = external_id[-6:] if len(external_id) >= 6 else external_id
            customer_name = f"{platform_labels.get(platform, platform)} #{suffix}"
        
        platform_labels = {'grabfood': 'GrabFood', 'gofood': 'GoFood', 'shopeefood': 'ShopeeFood'}
        order = Order(
            order_number=order_number,
            customer_name=customer_name,
            order_type='online',
            source=platform,
            status='processing',
            notes=f"[{platform_labels.get(platform, platform)}] {notes}".strip(),
            branch_id=branch_id
        )
        db.session.add(order)
        db.session.flush()
        
        # Process items with price snapshot and menu matching
        total = 0
        for item in items_data:
            item_name = item.get('name', 'Item')
            item_qty = item.get('quantity', 1)
            item_price = item.get('price', 0)
            item_notes = item.get('notes', '')
            item_ext_id = item.get('external_id', '')
            
            # Match menu item: try external_id first (partner's ID), then exact name, then partial
            menu_item = None
            if item_ext_id:
                menu_item = MenuItem.query.filter_by(id=item_ext_id).first() if isinstance(item_ext_id, str) and item_ext_id.isdigit() else None
            if not menu_item:
                menu_item = MenuItem.query.filter(
                    db.func.lower(MenuItem.name) == item_name.lower()
                ).first()
            if not menu_item:
                menu_item = MenuItem.query.filter(
                    MenuItem.name.ilike(f'%{item_name}%')
                ).first()
            
            menu_price_snapshot = menu_item.price if menu_item else None
            if menu_item and item_price == 0:
                item_price = menu_item.price
            
            subtotal = item_price * item_qty
            total += subtotal
            
            order_item = OrderItem(
                order_id=order.id,
                menu_item_id=menu_item.id if menu_item else None,
                name=item_name,
                price=item_price,
                menu_price_snapshot=menu_price_snapshot,
                quantity=item_qty,
                subtotal=subtotal,
                notes=item_notes,
                item_status='pending'
            )
            db.session.add(order_item)
        
        # Use platform-provided total if available and our calculated total is 0
        if total == 0 and parsed.get('total', 0) > 0:
            total = parsed['total']
        
        order.subtotal = total
        order.total = total
        
        # Payment (pre-paid via platform)
        payment = Payment(
            order_id=order.id,
            payment_method=platform,
            amount=total,
            paid_amount=total,
            status='paid',
            paid_at=utc_now()
        )
        db.session.add(payment)
        
        # Update external order status â†’ accepted
        ext_order.order_id = order.id
        ext_order.status = 'accepted'
        ext_order.processed_at = utc_now()
        ext_order.accepted_at = utc_now()
        
        # Notification
        total_formatted = f"{total:,}".replace(',', '.')
        create_notification(
            type='order_new',
            title=f'Pesanan {platform_labels.get(platform, platform)}!',
            message=f'Order #{order_number} - {customer_name} - Rp {total_formatted}',
            data={'order_id': order.id, 'order_number': order_number, 'platform': platform}
        )
        
        # Schedule print (DEFERRED â€” NOT printed inline in webhook)
        receipt_data = [
            {"type": "text", "value": "================================", "align": "center"},
            {"type": "text", "value": platform_labels.get(platform, platform).upper(), "align": "center", "bold": True, "size": "large"},
            {"type": "text", "value": "================================", "align": "center"},
            {"type": "text", "value": f"Order: #{order_number}", "bold": True},
            {"type": "text", "value": f"Pelanggan: {customer_name}"},
            {"type": "text", "value": f"Waktu: {datetime.now().strftime('%d/%m/%Y %H:%M')}"},
            {"type": "text", "value": "--------------------------------", "align": "center"},
        ]
        
        for item in items_data:
            receipt_data.append({"type": "text", "value": f"{item.get('quantity', 1)}x {item.get('name', 'Item')}"})
            if item.get('price', 0) > 0:
                price_fmt = f"Rp {item['price'] * item.get('quantity', 1):,}".replace(',', '.')
                receipt_data.append({"type": "text", "value": f"   {price_fmt}", "align": "right"})
            if item.get('notes'):
                receipt_data.append({"type": "text", "value": f"   Catatan: {item['notes']}"})
        
        receipt_data.extend([
            {"type": "text", "value": "--------------------------------", "align": "center"},
            {"type": "text", "value": f"TOTAL: Rp {total:,}".replace(',', '.'), "bold": True, "align": "right"},
            {"type": "text", "value": "================================", "align": "center"},
        ])
        if notes:
            receipt_data.append({"type": "text", "value": f"Catatan: {notes}"})
        receipt_data.append({"type": "text", "value": ""})
        receipt_data.append({"type": "cut"})
        
        pending_print = PendingPrint(
            order_id=order.id,
            receipt_data=json.dumps(receipt_data),
            copies=2,
            status='pending',
            branch_id=branch_id
        )
        db.session.add(pending_print)
        
        db.session.commit()
        return ext_order, None
        
    except Exception as e:
        ext_order.status = 'failed'
        ext_order.error_message = str(e)
        db.session.commit()
        return ext_order, str(e)


# ---- Unified Webhook Handler ----

def handle_webhook(platform):
    """
    Unified webhook handler. Fast response, deferred printing.
    1. Verify platform-specific signature
    2. Log the attempt
    3. Create order (print scheduled, not inline)
    4. Return 200 quickly (platforms retry on timeout)
    """
    is_valid, sig_status, sig_detail = verify_webhook_signature(platform, request)
    
    if not is_valid:
        log_webhook_attempt(platform, request, 'sig_' + sig_status, sig_detail)
        return jsonify({'error': 'Unauthorized', 'detail': sig_detail}), 401
    
    data = request.get_json(silent=True)
    if not data:
        log_webhook_attempt(platform, request, 'error', 'Invalid JSON payload')
        return jsonify({'error': 'Invalid JSON payload'}), 400
    
    # Extract using platform-specific parser for branch resolution
    extractors = {
        'grabfood': extract_grabfood_order,
        'gofood': extract_gofood_order,
        'shopeefood': extract_shopeefood_order,
    }
    parsed = extractors[platform](data)
    branch_id = resolve_branch_from_platform(platform, parsed, request)
    
    ext_order, error = process_platform_order(platform, data, branch_id)
    
    if ext_order:
        ext_order.signature_status = sig_status
        db.session.commit()
    
    if error and not ext_order:
        log_webhook_attempt(platform, request, 'error', error)
        return jsonify({'success': False, 'error': error}), 400
    
    if error:
        log_webhook_attempt(platform, request, 'success', error, ext_order.id)
        return jsonify({'success': True, 'message': error, 'external_order_id': ext_order.id}), 200
    
    log_webhook_attempt(platform, request, 'success', None, ext_order.id)
    return jsonify({
        'success': True,
        'message': 'Order received',
        'order_id': ext_order.order_id,
        'external_order_id': ext_order.id
    }), 200


# ---- Webhook Endpoints ----

@csrf.exempt
@app.route('/api/webhook/grabfood', methods=['POST'])
@limiter.limit("60 per minute")
def webhook_grabfood():
    """GrabFood webhook â€” receives order via X-Grab-Signature (HMAC-SHA256)"""
    return handle_webhook('grabfood')


@csrf.exempt
@app.route('/api/webhook/gofood', methods=['POST'])
@limiter.limit("60 per minute")
def webhook_gofood():
    """GoFood webhook â€” receives order via X-Callback-Token"""
    return handle_webhook('gofood')


@csrf.exempt
@app.route('/api/webhook/shopeefood', methods=['POST'])
@limiter.limit("60 per minute")
def webhook_shopeefood():
    """ShopeeFood webhook â€” receives order via HMAC-SHA256 signature"""
    return handle_webhook('shopeefood')


# ---- Order Status Management ----

@csrf.exempt
@app.route('/api/webhook/order/<int:ext_id>/mark-printed', methods=['POST'])
@limiter.limit("60 per minute")
def webhook_mark_printed(ext_id):
    """Mark external order as printed (called by print worker after successful print)"""
    ext_order = ExternalOrder.query.get_or_404(ext_id)
    if ext_order.status == 'accepted':
        ext_order.status = 'printed'
        ext_order.printed_at = utc_now()
        db.session.commit()
    return jsonify({'success': True, 'status': ext_order.status}), 200


# ---- Platform API: Accept Order ----

@app.route('/api/platform/<platform>/accept/<int:ext_id>', methods=['POST'])
@login_required
@role_required('admin', 'kasir')
def platform_accept_order(platform, ext_id):
    """
    Accept an order on the platform side (for platforms that require manual acceptance).
    GrabFood: PUT /partner/v1/order/prepare
    GoFood: PUT /integrations/gofood/outlets/{outlet}/v1/orders/delivery/{order_id}/accepted
    """
    import requests as http_requests
    
    ext_order = ExternalOrder.query.get_or_404(ext_id)
    cfg = get_platform_config(platform)
    
    if platform == 'grabfood':
        token = get_oauth_token('grabfood')
        if not token:
            return jsonify({'success': False, 'error': 'Gagal mendapatkan token OAuth2 GrabFood. Cek CLIENT_ID/SECRET.'}), 400
        
        try:
            api_base = cfg['api_base']
            order_id = ext_order.external_order_id
            resp = http_requests.put(
                f'{api_base}/partner/v1/order/prepare',
                json={'orderID': order_id},
                headers={
                    'Authorization': f'Bearer {token}',
                    'Content-Type': 'application/json',
                },
                timeout=15
            )
            if resp.status_code in (200, 204):
                return jsonify({'success': True, 'message': 'Order diterima di GrabFood'})
            return jsonify({'success': False, 'error': f'GrabFood API error: {resp.status_code} {resp.text}'}), 400
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif platform == 'gofood':
        token = get_oauth_token('gofood')
        if not token:
            return jsonify({'success': False, 'error': 'Gagal mendapatkan token OAuth2 GoFood. Cek CLIENT_ID/SECRET.'}), 400
        
        try:
            api_base = cfg['api_base']
            outlet_id = cfg.get('merchant_id', '')
            order_id = ext_order.external_order_id
            resp = http_requests.put(
                f'{api_base}/integrations/gofood/outlets/{outlet_id}/v1/orders/delivery/{order_id}/accepted',
                headers={
                    'Authorization': f'Bearer {token}',
                    'Content-Type': 'application/json',
                },
                timeout=15
            )
            if resp.status_code in (200, 204):
                return jsonify({'success': True, 'message': 'Order diterima di GoFood'})
            return jsonify({'success': False, 'error': f'GoFood API error: {resp.status_code} {resp.text}'}), 400
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    return jsonify({'success': False, 'error': 'Platform tidak mendukung accept API'}), 400


# ---- Webhook Replay Tool (Admin) ----

@app.route('/admin/webhook/<int:ext_id>/replay', methods=['POST'])
@login_required
@role_required('admin')
def webhook_replay(ext_id):
    """Replay a failed/received webhook order from stored raw_data"""
    ext_order = ExternalOrder.query.get_or_404(ext_id)
    
    if ext_order.status not in ('failed', 'received'):
        return jsonify({'success': False, 'error': 'Hanya order gagal/received yang bisa di-replay'}), 400
    
    try:
        data = json.loads(ext_order.raw_data)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'success': False, 'error': 'Raw data tidak valid'}), 400
    
    old_platform = ext_order.platform
    branch_id = ext_order.branch_id
    retry_count = (ext_order.retry_count or 0) + 1
    
    # Remove old record so duplicate check allows re-processing
    db.session.delete(ext_order)
    db.session.commit()
    
    new_ext, error = process_platform_order(old_platform, data, branch_id)
    
    if new_ext:
        new_ext.retry_count = retry_count
        new_ext.last_retry_at = utc_now()
        db.session.commit()
    
    if error and not new_ext:
        return jsonify({'success': False, 'error': error}), 400
    
    return jsonify({
        'success': True,
        'message': 'Order berhasil di-replay',
        'order_id': new_ext.order_id if new_ext else None
    }), 200


# ---- Admin Integration Dashboard ----

@app.route('/admin/integrations')
@login_required
@role_required('admin')
def admin_integrations():
    """Food delivery platform integration settings and monitoring"""
    platforms = {}
    for key, pcfg in PLATFORM_CONFIG.items():
        cfg = get_platform_config(key)
        platforms[key] = {
            'name': pcfg['name'],
            'icon': {'grabfood': 'fa-motorcycle', 'gofood': 'fa-utensils', 'shopeefood': 'fa-shopping-bag'}[key],
            'color': {'grabfood': 'green', 'gofood': 'red', 'shopeefood': 'orange'}[key],
            'webhook_url': request.url_root.rstrip('/') + f'/api/webhook/{key}',
            'secret': cfg.get('webhook_secret', ''),
            'client_id': cfg.get('client_id', ''),
            'store_id': cfg.get('store_id', ''),
            'store_branch_map': cfg.get('store_branch_map', ''),
            'enabled': bool(cfg.get('webhook_secret') or cfg.get('client_id')),
            'sig_header': pcfg['sig_header'],
            'sig_algo': {'sha256': 'HMAC-SHA256', 'token': 'Token Comparison'}.get(pcfg['sig_algo'], pcfg['sig_algo']),
            'has_oauth': bool(pcfg.get('oauth_token_url')),
            'is_sandbox': cfg.get('is_sandbox', True),
        }
    
    recent_orders = ExternalOrder.query.order_by(ExternalOrder.created_at.desc()).limit(20).all()
    recent_logs = WebhookLog.query.order_by(WebhookLog.created_at.desc()).limit(20).all()
    
    from sqlalchemy import func
    today = datetime.now().date()
    today_stats = db.session.query(
        ExternalOrder.platform,
        func.count(ExternalOrder.id).label('count')
    ).filter(
        func.date(ExternalOrder.created_at) == today
    ).group_by(ExternalOrder.platform).all()
    stats = {s.platform: s.count for s in today_stats}
    
    yesterday = datetime.now() - timedelta(hours=24)
    sig_failures = WebhookLog.query.filter(
        WebhookLog.created_at >= yesterday,
        WebhookLog.result.in_(['sig_invalid', 'sig_missing'])
    ).count()
    
    return render_template('admin/integrations.html',
                         platforms=platforms,
                         recent_orders=recent_orders,
                         recent_logs=recent_logs,
                         stats=stats,
                         sig_failures=sig_failures,
                         active_page='admin_integrations')


# ========== PAYMENT GATEWAY ADMIN ==========
@app.route('/admin/payment-gateway')
@login_required
@role_required('admin')
def admin_payment_gateway():
    """Payment gateway settings page"""
    # Get current settings
    server_key = app.config.get('MIDTRANS_SERVER_KEY', '')
    client_key = app.config.get('MIDTRANS_CLIENT_KEY', '')
    is_production = app.config.get('MIDTRANS_IS_PRODUCTION', False)
    
    # Mask keys for display
    masked_server_key = server_key[:8] + '****' + server_key[-4:] if len(server_key) > 12 else '****'
    masked_client_key = client_key[:8] + '****' + client_key[-4:] if len(client_key) > 12 else '****'
    
    return render_template('admin/payment_gateway.html',
                         masked_server_key=masked_server_key,
                         masked_client_key=masked_client_key,
                         is_production=is_production,
                         is_configured=bool(server_key and client_key),
                         active_page='admin_payment_gateway')


@app.route('/api/payment-gateway/test', methods=['POST'])
@login_required
@role_required('admin')
def api_test_payment_gateway():
    """Test Midtrans API connection"""
    import base64
    import requests
    
    server_key = app.config.get('MIDTRANS_SERVER_KEY', '')
    is_production = app.config.get('MIDTRANS_IS_PRODUCTION', False)
    
    if not server_key:
        return jsonify({
            'success': False,
            'message': 'Server Key belum dikonfigurasi'
        })
    
    # Choose the correct API URL
    if is_production:
        api_url = 'https://api.midtrans.com/v2/ping'
    else:
        api_url = 'https://api.sandbox.midtrans.com/v2/ping'
    
    try:
        # Create auth header
        auth_string = base64.b64encode(f'{server_key}:'.encode()).decode()
        headers = {
            'Authorization': f'Basic {auth_string}',
            'Accept': 'application/json'
        }
        
        response = requests.get(api_url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            return jsonify({
                'success': True,
                'message': f'Koneksi berhasil! Mode: {"Production" if is_production else "Sandbox"}',
                'environment': 'production' if is_production else 'sandbox'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Koneksi gagal: HTTP {response.status_code}'
            })
    except requests.exceptions.Timeout:
        return jsonify({
            'success': False,
            'message': 'Koneksi timeout. Coba lagi nanti.'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


if __name__ == '__main__':
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    os.makedirs('static/qrcodes', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    os.makedirs('uploads', exist_ok=True)
    
    # Initialize database
    init_db()
    
    print("""
    ðŸ½ï¸  KASIR MODERN - FULL FEATURES
    ====================================
    ðŸŽ¯ FITUR:
    1. âœ… Login & Register
    2. âœ… Role & Permission
    3. âœ… Pesanan Manual & Online (QR Code)
    4. âœ… Profile & Logout
    5. âœ… Payment Gateway (Midtrans)
    6. âœ… Spice Level & Hot/Cold Options
    7. âœ… Statistics & Reports (PDF & Excel)
    8. âœ… Admin Management
    9. âœ… Income Management
    10. âœ… Menu dari PDF Solaria
    11. âœ… Kitchen Display untuk Koki
    12. âœ… Manajemen Meja (Tambah/Hapus)
    
    ðŸ“± Modern UI dengan Tailwind CSS
    ðŸŽ¨ Glassmorphism Design
    ðŸ” Secure Authentication
    
    ðŸ‘¤ Default Login:
       Admin: admin / admin123
       Kasir: kasir / kasir123
       Koki:  koki / koki123
    
    ðŸŒ Server: http://localhost:8000
    """)
    
    # Use debug mode only in development (controlled by environment variable)
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=8000, use_reloader=debug_mode)
